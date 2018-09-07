# compare wv3 and sentinel rsr's for xcalib

import numpy as np
import pylab
import os
import scipy.stats
import matplotlib.pyplot as pyplot
import matplotlib as mpl


def ReadRsr(fileName, sheetName='Sheet1'):
    from openpyxl import load_workbook
    wb = load_workbook(fileName)
    sheet = wb[sheetName]
    wavelen = []
    rsr = []
    labels=[]
    # wavelen Pa B1	B2 B3
    for cell in sheet[1]:
        labels.append(cell.value)
    labels = labels[1:]
    for i in range(sheet.min_row+1, sheet.max_row):
        row = sheet[i]
        wavelen.append(row[0].value)
        # v = [0., 0., 0., 0.]
        # cols = [3,2,1] #nir, r, g - same as DMC ordering!
        v = np.zeros((sheet.max_column - sheet.min_column))
        for j in range(1, sheet.max_column-1):
            v[j-1] = row[j].value
            if v[j-1] is None:
                v[j-1] = 0.
        rsr.append(v)
    rsr = np.array(rsr)
    return wavelen, rsr, labels

wv3FileName = "C:/Data/Development/Projects/PhD GeoInformatics/Data/Spectral Sensitivities/WorldView3.xlsx"
sentinelFileName = "C:/Data/Development/Projects/PhD GeoInformatics/Data/Spectral Sensitivities/S2-SRF_COPE-GSEG-EOPG-TN-15-0007_3.0.xlsx"

wv3Wavelen,wv3Rsr, wv3Labels = ReadRsr(wv3FileName, 'Sheet1')
senWavelen,senRsr, senLabels = ReadRsr(sentinelFileName, 'Spectral Responses (S2A)')

pylab.figure()
pylab.plot(wv3Wavelen, wv3Rsr, '-')
pylab.legend(wv3Labels)

pylab.figure()
pylab.plot(senWavelen, senRsr, '-')
pylab.legend(senLabels)

pylab.figure()
pylab.plot(wv3Wavelen, wv3Rsr[:,1:9], '-')
pylab.plot(senWavelen, senRsr[:,:10], '--')
pylab.legend(wv3Labels[1:9] + senLabels[:10])

##############################################################################################
# So the mapping should look something like
#
# WV3                           Sentinel
# 1 Coastal: 400 - 450 nm       b1 Coastal aerosol
# 2 Blue: 450 - 510 nm          b2 Blue
# 3 Green: 510 - 580 nm         b3 Green
# 4 Yellow: 585 - 625 nm        ? No corresponding band -  b3 is closest but does not overlap
# 5 Red: 630 - 690 nm           b4 Red
# 6 Red Edge: 705 - 745 nm      (b5+b6)/2  (Vegetation red edge + Vegetation red edge)/2
# 7 Near-IR1: 770 - 895 nm      b8 NIR
# 8 Near-IR2: 860 - 1040 nm     (b8a + b9)/2 (Narrow NIR + Water Vapour)/2

