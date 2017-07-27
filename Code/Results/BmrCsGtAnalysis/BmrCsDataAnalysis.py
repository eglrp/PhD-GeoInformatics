################################################################################################
# Analyse planting dates etc for Baviaans CS GT
################################################################################################


import openpyxl
import gdal
import sys
import ogr
import numpy as np
import osr
import pylab
from scipy import stats as stats
from collections import OrderedDict

csGtFileName = "D:/Data/Development/Projects/MSc GeoInformatics/Data/Misc/BMR Carbon Stocks/abf_agc_191_plots.shp"
plantingFileName = "D:/Data/Development/Projects/MSc GeoInformatics/Data/Misc/BMR ShapeFiles/Baviaans_contract_Data.shp"
vegMapFileName = "D:/Data/Development/Projects/MSc GeoInformatics/Data/Misc/BMR ShapeFiles/vlokveg2010_Tm25.shp"


################################################################################################
# Read in MP's CS GT
################################################################################################

ds = gdal.OpenEx(csGtFileName, gdal.OF_VECTOR)
if ds is None:
    print "Open failed./n"

lyr = ds.GetLayerByIndex(0)
lyr.ResetReading()
csSpatialRef = lyr.GetSpatialRef()

csPlotList = []
for (i, feat) in enumerate(lyr):
    if i > 190:
        break
    print '.',
    feat_defn = lyr.GetLayerDefn()
    f = {}
    for i in range(feat_defn.GetFieldCount()):
        field_defn = feat_defn.GetFieldDefn(i)
        f[field_defn.GetName()] = feat.GetField(i)
    geom = feat.GetGeometryRef()
    if geom is not None and geom.GetGeometryType() == ogr.wkbPoint:
        # print "%.3f, %.3f" % ( geom.GetX(), geom.GetY() )
        f['geom'] = geom
        f['X'] = geom.GetX()
        f['Y'] = geom.GetY()
    else:
        print "no point geometry/n"
    f['Planted'] = False
    f['FirstPlantingYear'] = 0
    f['FirstPlantingMonth'] = ''
    f['LastPlantingYear'] = 0
    csPlotList.append(f)
print ' '
ds = None
geg = geom
csPlotList = csPlotList[:191]

################################################################################################
# Read GIB data and store dates and CSS ID's
################################################################################################
if True:
    gibFileName = "D:/Data/Development/Projects/MSc GeoInformatics/Data/Misc/Thicket Contract Summary 31 March 2016.xlsx"
    gibFileName = "C:/Data/Development/Projects/MSc GeoInformatics/Docs/Misc/Baviaanskloof/Thicket Contract Summary 31 March 2016.xlsx"

    from openpyxl import load_workbook
    wb = load_workbook(gibFileName)
    ws = wb["GIS Baviaans Summary March 2016"]

    first_row = ws[0+1]
    header=[]
    for c in first_row:
        header.append(c.value)

    months = []
    gibCssIds = []
    for r in ws[2:header.__len__()]:
        months.append(r[1].value)
        gibCssIds.append(r[-2].value)

    gibYears = np.int32([m[-4:] for m in months])
    gibCssIds = np.array(gibCssIds)
    gibMonths = [m[:-5] for m in months]

else:
    gibFileName = "F:/MSc GeoInformatics/Data/Misc/Plot_Codes Query.xlsx"

    from openpyxl import load_workbook
    wb = load_workbook(gibFileName)
    ws = wb["Plot_Codes Query"]

    first_row = ws.rows[0]
    header=[]
    for c in first_row:
        header.append(c.value)

    dates = []
    gibCssIds = []
    years = []
    for r in ws.rows[1:]:
        dates.append(r[-2].value)
        gibCssIds.append(r[2].value)
        if r[-2].value is not None:
            years.append(r[-2].value.year)
        else:
            years.append(0)

    gibYears = np.array(years)  # np.int32([d.year for d in dates])
    gibCssIds = np.array(gibCssIds)


################################################################################################
# Read CSS shapefile to determine which CS GT plots have been planted and at what date
################################################################################################

ds = gdal.OpenEx(plantingFileName, gdal.OF_VECTOR)
if ds is None:
    print "Open failed./n"

lyr = ds.GetLayerByIndex(0)
lyr.ResetReading()

plantingSpatialRef = lyr.GetSpatialRef()

# setup reprojection
# testing
print csSpatialRef
print plantingSpatialRef

coordTrans = osr.CoordinateTransformation(plantingSpatialRef, csSpatialRef)

plantingList = []
for feat in lyr:
    print '.',
    feat_defn = lyr.GetLayerDefn()
    f = {}
    for i in range(feat_defn.GetFieldCount()):
        field_defn = feat_defn.GetFieldDefn(i)
        f[field_defn.GetName()] = feat.GetField(i)
    geom = feat.GetGeometryRef()
    if geom is not None and geom.GetGeometryType() == ogr.wkbPolygon:
        # print "%.3f, %.3f" % ( geom.GetX(), geom.GetY() )
        geom.Transform(coordTrans)
        f['geom'] = geom
        b = geom.GetBoundary()
        if b.GetPointCount() > 0:
            pts = b.GetPoints()
            f['pts'] = pts
            f['X'] = [x for (x, y) in pts]
            f['Y'] = [y for (x, y) in pts]
        else:
            f['pts'] = []
            f['X'] = []
            f['Y'] = []

        f['FirstPlantingYear'] = 0
        f['LastPlantingYear'] = 0

        for plot in csPlotList:
            pt = ogr.Geometry(ogr.wkbPoint)
            pt.AddPoint_2D(plot['X'], plot['Y'])
            if geom.Contains(pt):
                # lookup planting year
                idx = gibCssIds == f['CSS_ID']
                years = np.array(gibYears[idx])
                months = np.array(gibMonths)[idx]
                if years.size > 0:
                    f['FirstPlantingMonth'] = months[years.argmin()]
                    f['FirstPlantingYear'] = years.min()
                    f['LastPlantingYear'] = years.max()
                    plot['Planted'] = True  # NB here we assume if this polygon is not in GIB data, then it was not plantes
                    plot['FirstPlantingMonth'] = months[years.argmin()]
                    plot['FirstPlantingYear'] = years.min()
                    plot['LastPlantingYear'] = years.max()
                    print 'x',
                else:
                    print 'Unknown CSS ID: ' + f['CSS_ID']

    plantingList.append(f)
print ' '
ds = None

################################################################################################
# Read VLok's veg map to get DST/ST/OL labels for MP CS GT with dodgy names
################################################################################################
ds = gdal.OpenEx(vegMapFileName, gdal.OF_VECTOR)
if ds is None:
    print "Open failed./n"

lyr = ds.GetLayerByIndex(0)
lyr.ResetReading()

for feat in lyr:
    print '.',
    feat_defn = lyr.GetLayerDefn()
    f = {}
    for i in range(feat_defn.GetFieldCount()):
        field_defn = feat_defn.GetFieldDefn(i)
        f[field_defn.GetName()] = feat.GetField(i)
    geom = feat.GetGeometryRef()
    if geom is not None and geom.GetGeometryType() == ogr.wkbPolygon:
        # print "%.3f, %.3f" % ( geom.GetX(), geom.GetY() )
        geom.Transform(coordTrans)

        for plot in csPlotList:
            if plot['PLOT'].find('OL') < 0 and plot['PLOT'].find('ST') < 0:  #if not labelled
                pt = ogr.Geometry(ogr.wkbPoint)
                pt.AddPoint_2D(plot['X'], plot['Y'])
                if geom.Contains(pt):
                    if f['DEGR'].find('Pristine') >= 0:
                        plot['PLOT'] += '_ST'
                    elif f['DEGR'].find('Land') >= 0:
                        plot['PLOT'] += '_OL'
                    else:
                        plot['PLOT'] += '_DST'
                    print 'x',
print ' '

# make class label
# (1, 2, 3) = (ST, DST, OL)
for plot in csPlotList:
    if plot['PLOT'].find('DST') >= 0:
        fclass = 2
    elif plot['PLOT'].find('OL') >= 0:
        fclass = 3
    elif plot['PLOT'].find('ST') >= 0:
        fclass = 1
    else:
        fclass = 3  # hack from inspection of map
    plot['CLASS'] = fclass
print ' '

ds = None

################################################################################################
# Data analysis
################################################################################################
plotNames = [x['PLOT'] for x in csPlotList]
planted = np.array([x['Planted'] for x in csPlotList])
tagc = np.array([x['TAGC'] for x in csPlotList])
firstPlantedMonth = np.array([x['FirstPlantingMonth'] for x in csPlotList])
firstPlanted = np.array([x['FirstPlantingYear'] for x in csPlotList])
lastPlanted = np.array([x['LastPlantingYear'] for x in csPlotList])
plotClass = np.array([x['CLASS'] for x in csPlotList])
zPafra = np.array([x['Z_P_AFRA'] for x in csPlotList])
zPafra[zPafra < 0] = 0  # get rid of Nones
cssId = [x['CSS_ID'] for x in plantingList]
classLabels = ['All', 'ST', 'DST', 'OL']


#
print 'Portion of planted plots with dates: %f' % (float((firstPlanted > 0).sum())/float(planted.sum()))

# we must assume worst case planted plots without dates were planted in 2005
firstPlanted[np.logical_and(planted, firstPlanted==0)] = 2005
lastPlanted[np.logical_and(planted, lastPlanted==0)] = 2005

planted = firstPlanted > 0
blanked = lastPlanted > firstPlanted

print 'Portion of plots planted: %f' % (float(planted.sum())/float(planted.size))
print 'Portion of plots blanked: %f' % (float(blanked.sum())/float(planted.size))
# (float(planted.sum())/float(planted.size)) - (float(blanked.sum())/float(planted.size))

# planting stats per year
plantingYears = np.unique(np.append(firstPlanted[planted], lastPlanted[planted]))
for y in plantingYears:
    print 'Portion of plots first planted in %d: %f' % (y, float((firstPlanted == y).sum())/float(firstPlanted.size))
    print 'Portion of plots blanked in %d: %f' % (y, float(np.logical_and(blanked, lastPlanted == y).sum())/float(lastPlanted.size))

# planting stats per year and degradation class
plantingStatListByDegrClass = []
for c in range(0, 4):
    if c == 0:  # all
        c_idx = np.bool_(np.ones_like(planted))
    else:
        c_idx = plotClass == c
    ps = OrderedDict()
    ps['Class'] = classLabels[c]
    ps['N'] = c_idx.sum()
    ps['Mean(TAGC)'] = np.mean(tagc[c_idx])
    ps['5th percentile(TAGC)'] = np.percentile(tagc[c_idx], 5)
    ps['Mean(Z_P_AFRA)'] = np.mean(zPafra[c_idx])
    ps['% Planted'] = 100*float(planted[c_idx].sum())/float(c_idx.sum())
    ps['% Blanked'] = 100*float(blanked[c_idx].sum())/float(c_idx.sum())
    for y in plantingYears:
        cy_idx = np.logical_and(firstPlanted == y, c_idx)
        ps['% Planted ' + str(y)] = 100*float(planted[cy_idx].sum())/float(c_idx.sum())
        cyb_idx = (lastPlanted == y) & (lastPlanted > firstPlanted) & c_idx
        ps['% Blanked ' + str(y)] = 100*float(blanked[cyb_idx].sum())/float(c_idx.sum())

    print classLabels[c]
    print ps
    plantingStatListByDegrClass.append(ps)


plantingStatListByYear = []

for y in plantingYears:
    ps = OrderedDict()
    ps['Year'] = y
    y_idx = firstPlanted == y
    ps['% Planted'] = 100*float(planted[y_idx].sum())/float(y_idx.size)
    yb_idx = np.logical_and(lastPlanted == y, lastPlanted > firstPlanted)
    ps['% Blanked'] = 100*float(blanked[yb_idx].sum())/float(y_idx.size)

    print y
    print ps
    plantingStatListByYear.append(ps)

ps = {}
ps['Year'] = 'All'
ps['% Planted'] = 100*float(planted.sum())/float(planted.size)
ps['% Blanked'] = 100*float(blanked.sum())/float(blanked.size)
plantingStatListByYear.append(ps)

# plotClass = np.zeros(plotNames.__len__())
# for (i, f) in enumerate(csPlotList):
#     if f['PLOT'].find('DST') >= 0:
#         plotClass[i] = 2
#     elif f['PLOT'].find('OL') >= 0:
#         plotClass[i] = 3
#     elif f['PLOT'].find('ST') >= 0:
#         plotClass[i] = 1
#     else:
#         plotClass[i] = 3  # hack from inspection of map

# Per degradation class stats and histogram
pylab.figure()
for i in range(1, 4):
    idx = plotClass == i
    # plantCount = planted[idx].sum()
    print 'Class: %s, count: %d, planted: %d, tagc median: %f, 5ptile: %f' % (classLabels[i], idx.sum(), planted[idx].sum(),
                                                                np.median(tagc[idx]), np.percentile(tagc[idx], 10))
    h, x = np.histogram(tagc[idx], density=True, bins=10)
    # pylab.subplot(2, 2, i+1)
    pylab.plot(x[:-1] + np.diff(x).mean()/2, h, '.-')
    pylab.xlabel('TAGC')
    pylab.hold('on')
pylab.legend(classLabels[1:])
pylab.grid()
pylab.title('Class TAGC histograms')

# Per year figures
pylab.figure()
h1, x1 = np.histogram(firstPlanted[firstPlanted>0], density=True, bins=range(2004, 2016))
pylab.plot(x1[:-1], h1, '.-')
pylab.hold('on')
h2, x2 = np.histogram(lastPlanted[lastPlanted>0], density=True, bins=range(2004, 2016))
pylab.plot(x2[:-1], h2, '.-')
pylab.legend(['first', 'last'])
pylab.grid()
pylab.xlabel('Year')
pylab.ylabel('Portion planted')
pylab.title('Planting year histogram')

pylab.figure()
pylab.plot(x1[:-1], np.cumsum(h1), '.-')
pylab.hold('on')
pylab.plot(x2[:-1], np.cumsum(h2), '.-')
pylab.legend(['first', 'last'])
pylab.grid()
pylab.xlabel('Year')
pylab.ylabel('Portion planted')
pylab.title('Planting year cumulative histogram')


##########################################################################################
# write to file
##########################################################################################
from csv import DictWriter
with open('C:\Data\Development\Projects\MSc GeoInformatics\Docs\Misc\Baviaanskloof\PlantingStatsDegrClass.csv','w') as outfile:
    writer = DictWriter(outfile, plantingStatListByDegrClass[0].keys())
    writer.writeheader()
    writer.writerows(plantingStatListByDegrClass)

with open('C:\Data\Development\Projects\MSc GeoInformatics\Docs\Misc\Baviaanskloof\PlantingStatsYear.csv','w') as outfile:
    writer = DictWriter(outfile, plantingStatListByYear[0].keys())
    writer.writeheader()
    writer.writerows(plantingStatListByYear)



################################################################################################
# Further analysis of Quickbird image coverage
################################################################################################

qbFileNames = ["C:/Data/Development/Projects/MSc GeoInformatics/Docs/Misc/Baviaanskloof/QuickBird 2003_11_18 10100100027B8601.shp",
               "C:/Data/Development/Projects/MSc GeoInformatics/Docs/Misc/Baviaanskloof/QuickBird 2005_06_06 10100100044A0801.shp"]
qbCodes = ["QB2003", "QB2005"]

for (qbFileName, qbCode) in zip(qbFileNames, qbCodes):
    ds = gdal.OpenEx(qbFileName, gdal.OF_VECTOR)
    if ds is None:
        print "Open failed./n"

    lyr = ds.GetLayerByIndex(0)
    lyr.ResetReading()

    for feat in lyr:
        print '.',
        feat_defn = lyr.GetLayerDefn()
        f = {}
        for i in range(feat_defn.GetFieldCount()):
            field_defn = feat_defn.GetFieldDefn(i)
            f[field_defn.GetName()] = feat.GetField(i)
        geom = feat.GetGeometryRef()
        if geom is not None and geom.GetGeometryType() == ogr.wkbPolygon:
            # print "%.3f, %.3f" % ( geom.GetX(), geom.GetY() )
            # geom.Transform(coordTrans)

            for plot in csPlotList:
                pt = ogr.Geometry(ogr.wkbPoint)
                pt.AddPoint_2D(plot['X'], plot['Y'])
                if geom.Contains(pt):
                    plot[qbCode] = True
                    print 'x',
                else:
                    plot[qbCode] = False
                    print '-',
                # if plot['PLOT'].find('OL') < 0 and plot['PLOT'].find('ST') < 0:  #if not labelled
                #     pt = ogr.Geometry(ogr.wkbPoint)
                #     pt.AddPoint_2D(plot['X'], plot['Y'])
                #     if geom.Contains(pt):
                #         if f['DEGR'].find('Pristine') >= 0:
                #             plot['PLOT'] += '_ST'
                #         elif f['DEGR'].find('Land') >= 0:
                #             plot['PLOT'] += '_OL'
                #         else:
                #             plot['PLOT'] += '_DST'
                #         print 'x',
    print ' '


################################################################################################
# Data analysis
################################################################################################
# plotNames = [x['PLOT'] for x in csPlotList]
# planted = np.array([x['Planted'] for x in csPlotList])
# tagc = np.array([x['TAGC'] for x in csPlotList])
# firstPlantedMonth = np.array([x['FirstPlantingMonth'] for x in csPlotList])
# firstPlanted = np.array([x['FirstPlantingYear'] for x in csPlotList])
# lastPlanted = np.array([x['LastPlantingYear'] for x in csPlotList])
# plotClass = np.array([x['CLASS'] for x in csPlotList])
# zPafra = np.array([x['Z_P_AFRA'] for x in csPlotList])
# zPafra[zPafra < 0] = 0  # get rid of Nones
# cssId = [x['CSS_ID'] for x in plantingList]
# classLabels = ['All', 'ST', 'DST', 'OL']


# plantingYears = np.unique(np.append(firstPlanted[planted], lastPlanted[planted]))
for qbCode in qbCodes:
    qbIntersect = np.array([x[qbCode] for x in csPlotList])
    print 'Portion of plots in %s: %f' % (qbCode, float(np.int32(qbIntersect).sum())/float(qbIntersect.size))

    # portion per class
    for c in range(0, 4):
        if c == 0:  # all
            c_idx = np.bool_(np.ones_like(csPlotList))
        else:
            c_idx = plotClass == c
        c_qb_idx = np.logical_and(c_idx, qbIntersect)
        ps = OrderedDict()
        ps['Im'] = qbCode
        ps['Class'] = classLabels[c]
        ps['N'] = c_qb_idx.sum()
        ps['Portion'] = float(np.int32(c_qb_idx).sum())/float(qbIntersect.size)

        print 'Portion, num of %s plots in %s: %f, %d' % (classLabels[c], qbCode, float(np.int32(c_qb_idx).sum())/float(qbIntersect.size), np.int32(c_qb_idx).sum())

