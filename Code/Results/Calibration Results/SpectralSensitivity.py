# %gui wx
# %matplotlib wx

# from SpectralSensitivity import asterRefl

import numpy as np
import pylab
import os
import scipy.stats
import matplotlib.pyplot as pyplot
import matplotlib as mpl

def ReadLandsatRsr(fileName):
    from openpyxl import load_workbook
    wb = load_workbook(fileName)
    sheetNames = ['Blue-L7', 'Green-L7', 'Red-L7', 'NIR-L7']
    wavelenList = []
    rsrList = []
    for sheetName in sheetNames:
        sheet = wb[sheetName]
        wavelen = []
        rsr = []
        for row in sheet.iter_rows('A{}:B{}'.format(sheet.min_row+1, sheet.max_row)):
            wavelen.append(row[0].value)
            rsr.append(row[1].value)
        wavelenList.append(wavelen)
        rsrList.append(rsr)
    return wavelenList, rsrList


def ReadModisRsr(filename, ch = 5):
    t = np.loadtxt(filename, dtype = float, comments = '#')
    #wavelen = t[:,2]
    #rsr = t[:,3]
    #ch = t[:,0]
    #r = None
    #for i in range(1, np.int(t[:,1].max()+1)):
    #    idx = (t[:,1] == i) & (t[:,3]>=0)
    #    if r is None:
    #        r = t[idx,3]
    #    else:
    #        r = r + t[idx,3]
    #rsr = r/ t[:,1].max()
    #NB only taking the first "channel" (= one of many measurements of a band?? at not exactly the same wavelens???)
    idx = (t[:,1] == ch) & (t[:,3]>=0)
    wavelen = t[idx,2]
    rsr = t[idx,3]
    return (wavelen, rsr)

def ReadAsterSpecLib(filename):
    t = np.loadtxt(filename, dtype = float, comments = '#', skiprows = 26)
    wavelen = t[:,0]
    refl = t[:,1]
    return (wavelen, refl)

def ReadDmcRsr(filename):
    t = np.loadtxt(filename, dtype = float, skiprows = 1) #, delimiter = ",")
    #Lambda	PAN	NIR	Red	Green	Blue
    wavelen = t[:,0]
    rsr = t[:,2:] 
    return (wavelen, rsr)

def ReadQuickbirdRsr(filename):
    t = np.loadtxt(filename, dtype = float, skiprows = 1) #, delimiter = ",")
    #Lambda	PAN	NIR	Red	Green	Blue
    wavelen = t[:,0]
    rsr = np.fliplr(t[:,2:])
    return (wavelen, rsr)

def SimSensorMeasurement(sensorRsr, surfRefl):
    sensorMeas = []
    for i in range(0, np.shape(sensorRsr)[1]):
        # normalise the RSR so that band avg refl is max 1 (this is a scaling thing to make reviewers happy for the paper)
        sensorRsrNorm = sensorRsr[:,i]/sensorRsr[:,i].sum()
        sensorMeas.append(np.inner(sensorRsrNorm, surfRefl))
    return sensorMeas

def ReadSpot5Rsr(fileName):
    from openpyxl import load_workbook
    wb = load_workbook(fileName)
    sheet = wb['SPOT 5']
    wavelen = []
    rsr = []
    # wavelen Pa B1	B2 B3

    for i in range(8, 108):
        row = sheet[i]
        wavelen.append(row[1].value)
        v = [0., 0., 0., 0.]
        cols = [3,2,1] #nir, r, g - same as DMC ordering!
        for j in range(3):
            v[j] = row[cols[j]+2].value
            if v[j] is None:
                v[j] = 0.
        rsr.append(v)
    return wavelen, rsr


dmcFn = "C:/Data/Development/Projects/PhD GeoInformatics/Data/Spectral Sensitivities/DMC Spectral sensitivity.txt" #ordered nir,r,g,b/
qbFn = "C:/Data/Development/Projects/PhD GeoInformatics/Data/Spectral Sensitivities/Quickbird Spectral Sensitivity.txt" #ordered nir,r,g,b

modisFn = ["C:/Data/Development/Projects/PhD GeoInformatics/Data/Spectral Sensitivities/rsr.2.inb.final",
           "C:/Data/Development/Projects/PhD GeoInformatics/Data/Spectral Sensitivities/rsr.1.inb.final",
           "C:/Data/Development/Projects/PhD GeoInformatics/Data/Spectral Sensitivities/rsr.4.inb.final",
           "C:/Data/Development/Projects/PhD GeoInformatics/Data/Spectral Sensitivities/rsr.3.inb.final"] #ordered here to be nir,r,g,b

spotFn = "C:/Data/Development/Projects/PhD GeoInformatics/Data/Spectral Sensitivities/SPOT r453_9_spectralsensivity.xlsx"

asterDir = "C:/Data/Development/Projects/PhD GeoInformatics/Data/Spectral Sensitivities"

landsatFn = "C:/Data/Development/Projects/PhD GeoInformatics/Data/Spectral Sensitivities/L7_RSR.xlsx"


# Note on units:
#
# -The raw aster spectra are surf refl in % - the interp'd ones are refl 0-1
# -The MODIS RSR's are max 1 - have they been normalised this way?
# -The DMC RSR's are max < 1 and wider than the MODIS RSR
# -The SPOT RSR's are max 1
# -SimSensorMeasurement just sums surf refl * rsr (inner product) it is not actual integration taking wavelen gap into
# account.  But if it is a simple "rectangular" integration, the dlambda is constant for numerator and denominator and
# thus can be factored out of the integral and num and den the cancel each other out.  So I think SimSensorMeasurement
# is actually correct.
# -The output of SimSensorMeasurement should be a band averaged reflectance but we get >1 values for both dmc and modis...
# -In other papers, the RSR is max 1
# -For the band averaged value to be max 1 however, I think it is necessary to normalise the RSR by its sum.  As the
# surface reflectance can be 1 across all wavelen, the max of the band avg vals will be the sum of the RSR...



modisWaveLen = []
modisRsr = []
pylab.figure('MODIS')
for mf in modisFn:
    (lm,rsr) = ReadModisRsr(mf, 1)
    modisWaveLen.append(lm)
    modisRsr.append(rsr)
    pylab.hold(True)
    pylab.plot(lm, rsr, '-')
    print(mf)

## full set of aster spectra
asterWaveLenF = []
asterReflF = []
asterFnF = []
asterHF = []
pylab.figure('ASTERF')
for afn in os.listdir(asterDir):
    if afn.endswith(".spectrum.txt"):
        (al,ar) = ReadAsterSpecLib(asterDir + '/' + afn)
        if al.min() < 0.5: #only include visible range spectra
            asterWaveLenF.append(al)
            asterReflF.append(ar)
            afnSplit = afn.split('.')
            asterFnF.append('%s.%s.%s' % tuple(afnSplit[2:5]))
            pylab.hold(True)
            hl = pylab.plot(al, ar, '-')
            asterHF.append(hl[0])

#hAster = pylab.plot(np.asarray(asterWaveLen).transpose(), np.asarray(asterRefl).transpose(), '-')
pylab.legend(np.asarray(asterHF), asterFnF)

## subset of aster spectra for plotting
asterDir2 = "D:/Data/Development/Projects/PhD GeoInformatics/Data/Spectral Sensitivities/Temp"

asterWaveLen = []
asterRefl = []
asterFn = []
asterH = []
pylab.figure('ASTER')
for afn in os.listdir(asterDir2):
    if afn.endswith(".myspectrum.txt"):
        (al,ar) = ReadAsterSpecLib(asterDir2 + '/' + afn)
        if al.min() < 0.5: #only include visible range spectra
            asterWaveLen.append(al)
            asterRefl.append(ar)
            sf = afn.split('.')
            asterFn.append(sf[-3])
            pylab.hold(True)
            hl = pylab.plot(al, ar, '-')
            asterH.append(hl[0])

#hAster = pylab.plot(np.asarray(asterWaveLen).transpose(), np.asarray(asterRefl).transpose(), '-')
pylab.legend(np.asarray(asterH), asterFn)

(dmcWaveLen,dmcRsr) = ReadDmcRsr(dmcFn)
pylab.figure('DMC')
pylab.plot(dmcWaveLen, dmcRsr, '-')

(qbWaveLen,qbRsr) = ReadQuickbirdRsr(qbFn)
pylab.figure('QB')
pylab.plot(qbWaveLen, qbRsr, '-')

(spotWaveLen,spotRsr) = ReadSpot5Rsr(spotFn)
pylab.figure('SPOT 5')
pylab.plot(spotWaveLen, spotRsr, '-')
pylab.legend(['B3', 'B2', 'B1'])


(landsatWaveLenList,landsatRsrList) = ReadLandsatRsr(landsatFn)
pylab.figure('Landsat')
for (wavelen, rsr) in zip(landsatWaveLenList,landsatRsrList):
    pylab.plot(wavelen, rsr, '-')
# pylab.legend(['B3', 'B2', 'B1'])


##
# interp all to dmc wavelen
pylab.figure('Interp')
hDmc = pylab.plot(dmcWaveLen, dmcRsr, 'k-')

modisRsrInterp = []
for (wl,rsr) in zip(modisWaveLen, modisRsr):
    rsrInterp = np.interp(dmcWaveLen, wl, rsr, left=0, right=0)
    modisRsrInterp.append(rsrInterp)

hModis = pylab.plot(dmcWaveLen, np.asarray(modisRsrInterp).transpose(), 'r-')


asterReflInterp = []
for (wl,refl) in zip(asterWaveLen, asterRefl):
    idx = np.argsort(wl)
    reflInterp = np.interp(dmcWaveLen, wl[idx]*1000.0, refl[idx]/100, left=0, right=0)
    asterReflInterp.append(reflInterp)
hAster = pylab.plot(dmcWaveLen, np.asarray(asterReflInterp).transpose(), 'b-')

asterReflInterpF = []
for (wl,refl) in zip(asterWaveLenF, asterReflF):
    idx = np.argsort(wl)
    reflInterp = np.interp(dmcWaveLen, wl[idx]*1000.0, refl[idx]/100, left=0, right=0)
    asterReflInterpF.append(reflInterp)

hAster = pylab.plot(dmcWaveLen, np.asarray(asterReflInterpF).transpose(), 'b-')
#pylab.legend([hDmc, hModis, hAster], ['DMC','MODIS','ASTER'])
#pylab.legend(hAster, asterFn)

spotRsrInterp = []
for i in range(3):
    rsrInterp = np.interp(dmcWaveLen, spotWaveLen, np.array(spotRsr)[:,i], left=0, right=0)
    spotRsrInterp.append(rsrInterp)

pylab.figure('SPOT 5 Interp')
hSpot = pylab.plot(dmcWaveLen, np.asarray(spotRsrInterp).transpose(), 'r-')


pylab.figure('ASTER_')
hAsterF = pylab.plot(dmcWaveLen, np.asarray(asterReflInterpF).transpose())
#pylab.legend([hDmc, hModis, hAster], ['DMC','MODIS','ASTER'])
pylab.legend(hAsterF, asterFnF)


qbRsrInterp = []
for i in range(3):
    rsrInterp = np.interp(dmcWaveLen, qbWaveLen, np.array(qbRsr)[:,i], left=0, right=0)
    qbRsrInterp.append(rsrInterp)

pylab.figure('QB Interp')
hQb = pylab.plot(dmcWaveLen, np.asarray(qbRsrInterp).transpose(), 'r-')

landsatRsrInterp = []
for (wavelen, rsr) in zip(landsatWaveLenList,landsatRsrList):
    rsrInterp = np.interp(dmcWaveLen, wavelen, np.array(rsr), left=0, right=0)
    landsatRsrInterp.append(rsrInterp)

pylab.figure('Landsat Interp')
hQb = pylab.plot(dmcWaveLen, np.asarray(landsatRsrInterp).transpose(), 'r-')



dmcMeas = []
modisMeas = []
spotMeas = []
qbMeas = []
for asterR in asterReflInterpF:
    dmcMeas.append(SimSensorMeasurement(dmcRsr, asterR))
    modisMeas.append(SimSensorMeasurement(np.asarray(modisRsrInterp).transpose(), asterR))
    spotMeas.append(SimSensorMeasurement(np.asarray(spotRsrInterp).transpose(), asterR))
    qbMeas.append(SimSensorMeasurement(np.asarray(qbRsrInterp).transpose(), asterR))

# figures for paper of linear rel betw real world meas
pylab.close('all')
fontSize = 12.
mpl.rcParams.update({'font.size': fontSize})

f1=pylab.figure('MODIS vs DMC')
f1.set_size_inches(6, 4.5, forward=True)
f2=pylab.figure('MODIS vs DMC 2')
f2.set_size_inches(6, 4.5, forward=True)
colors = ['orange', 'red', 'green', 'blue']
legend = ['NIR', 'Red', 'Green', 'Blue']
markers = ['v', 'o', '^', 's']
hf2=[]
for i in range(0, np.shape(modisMeas)[1]):
    x = np.asarray(modisMeas)[:,i]
    y = np.asarray(dmcMeas)[:,i]
    (slope, intercept, r, p, stde) = scipy.stats.linregress(x, y)
    pylab.figure('MODIS vs DMC')
    pylab.subplot(221+i)
    pylab.plot(x, y , 'kx', markersize=12.)
    pylab.hold('on')
    xi = np.linspace(x.min(), x.max(), 10)
    yi = slope*xi + intercept
    pylab.plot(xi, yi, 'k-')
    pylab.title(legend[i])
    pylab.text(xi.mean()*1.0, yi.mean()*0.6, '$R^2$ = ' + str.format('{0:.2f}', np.round(r**2, 2)), fontsize=fontSize+1)
    # pylab.xlabel(r'MODIS $\rho_t$')
    # pylab.ylabel(r'DMC $\rho_t$')
    pylab.xlabel(r'MODIS surface refl.')
    pylab.ylabel(r'DMC surface refl.')
    pylab.grid('on')
    # for j in range(asterFnF.__len__()):
    #     pylab.text(x[j]+.2, y[j], asterFnF[j])
    pylab.tight_layout()
    pyplot.locator_params(axis='y', nbins=5)#to specify number of ticks on both or any single axes
    pyplot.locator_params(axis='x', nbins=4)

    pylab.figure('MODIS vs DMC 2')
    hf2.append(pylab.plot(x, y, linestyle='', color=colors[i], marker=markers[i], markeredgecolor=colors[i],
                          label='_'+legend[i]))
    pylab.hold('on')
    pylab.plot(xi, yi, color=colors[i], linestyle='-', label=legend[i] + str.format(' - $R^2$ = {0:.2f}',
                                                                                    np.round(r**2, 2)))
    pylab.hold('on')
    pylab.tight_layout()
    pyplot.locator_params(axis='y', nbins=8)#to specify number of ticks on both or any single axes
    pyplot.locator_params(axis='x', nbins=8)
    #pylab.text(xi[-1]*1.05, yi[-1]*0.95, '$R^2$ = ' + str.format('{0:.2f}', r**2), color=colors[i])

f1.savefig('C:/Data/Development/Projects/PhD GeoInformatics/Docs/My Docs/Thesis/Retrieval of Surface Reflectance '
           'from Aerial Imagery/Figure 4 - DMC vs MODIS Band Averaged Relationship.eps', dpi=600)


pylab.figure('MODIS vs DMC 2')
pylab.title('MODIS vs DMC 2')
pylab.xlabel('MODIS')
pylab.ylabel('DMC')
pylab.grid('on')
pylab.legend(loc='upper left')


pylab.figure('DMC vs SPOT')
for i in range(0, np.shape(spotMeas)[1]):
    x = np.asarray(spotMeas)[:,i]
    y = np.asarray(dmcMeas)[:,i]
    (slope, intercept, r, p, stde) = scipy.stats.linregress(x, y)
    pylab.subplot(221 + i)
    pylab.plot(x, y , 'kx')
    pylab.hold('on')
    xi = np.linspace(x.min(), x.max(), 10)
    yi = slope*xi + intercept
    pylab.plot(xi, yi, 'k-')
    # pylab.title('Band ' + str(i) + ', R2 = ' + str(r**2) + str.format(' y = {0:.2f}x + {1:.2f}', slope, intercept))
    pylab.title(legend[i])
    pylab.text(x.mean()*1.0, y.mean()*0.6, '$R^2$ = ' + str.format('{0:.2f}', r**2) + str.format(' y = {0:.2f}x + {1:.2f}', slope, intercept), fontsize=fontSize+1)
    pylab.xlabel('SPOT5')
    pylab.ylabel('DMC')
    pylab.grid()
    # for j in range(asterFnF.__len__()):
    #     pylab.text(x[j]+0.2, y[j], asterFnF[j])

# figures for paper of linear rel betw real world meas
pylab.close('all')
fontSize = 24.
mpl.rcParams.update({'font.size': fontSize})

f1=pylab.figure('MODIS vs QB')
f1.set_size_inches(12, 9, forward=True)
colors = ['orange', 'red', 'green', 'blue']
legend = ['NIR', 'Red', 'Green', 'Blue']
markers = ['v', 'o', '^', 's']
hf2=[]
for i in range(0, np.shape(modisMeas)[1]):
    x = np.asarray(modisMeas)[:,i]
    y = np.asarray(qbMeas)[:,i]
    (slope, intercept, r, p, stde) = scipy.stats.linregress(x, y)
    pylab.figure('MODIS vs QB')
    pylab.subplot(221+i)
    pylab.plot(x, y , 'kx', markersize=12.)
    pylab.hold('on')
    xi = np.linspace(x.min(), x.max(), 10)
    yi = slope*xi + intercept
    pylab.plot(xi, yi, 'k-')
    pylab.title(legend[i])
    pylab.text(xi.mean()*1.0, yi.mean()*0.6, '$R^2$ = ' + str.format('{0:.2f}', np.round(r**2, 2)), fontsize=fontSize+1)
    # pylab.xlabel(r'MODIS $\rho_t$')
    # pylab.ylabel(r'DMC $\rho_t$')
    pylab.xlabel(r'MODIS surface refl.')
    pylab.ylabel(r'QB surface refl.')
    pylab.grid('on')
    # for j in range(asterFnF.__len__()):
    #     pylab.text(x[j]+.2, y[j], asterFnF[j])
    pylab.tight_layout()
    pyplot.locator_params(axis='y', nbins=5)#to specify number of ticks on both or any single axes
    pyplot.locator_params(axis='x', nbins=4)



########################################################################################################################
# MODIS, DMC RSR for paper
########################################################################################################################
fontSize = 12.
mpl.rcParams.update({'font.size': fontSize})

pylab.close('all')
f1 = pylab.figure('DMC Spectral Sensitivities')
f1.set_size_inches(6, 4.5, forward=True)
colors = ['k', 'r', 'g', 'b']
hModis = []
hDmc = []
for i in range(0, 4):
#    hModis.append(pylab.plot(modisWaveLen[i], modisRsr[i], color=colors[i], linestyle='-'))
    # hModis.append(pylab.plot(modisWaveLen[i], modisRsr[i], color='k', linestyle='-'))
    # pylab.hold('on')
    mask = dmcRsr[:, i] > 0.001
#    hDmc.append(pylab.plot(dmcWaveLen[mask], dmcRsr[mask, i], color=colors[i], linestyle='--'))
    hDmc.append(pylab.plot(dmcWaveLen[mask], dmcRsr[mask, i], color='k', linestyle='-'))
# pylab.legend((hModis[0][0], hDmc[0][0]), ('MODIS', 'DMC'), fontsize=fontSize-3.)
pylab.xlabel('Wavelength ($\mu m$)')
pylab.ylabel('Relative spectral response')
pylab.tight_layout()
f1.savefig('C:/Data/Development/Projects/PhD GeoInformatics/Docs/My Docs/Thesis/Retrieval of Surface Reflectance '
           'from Aerial Imagery/Figure - DMC RSR.eps', dpi=600)

f1 = pylab.figure('MODIS and DMC Spectral Sensitivities')
f1.set_size_inches(6, 4.5, forward=True)
colors = ['k', 'r', 'g', 'b']
hModis = []
hDmc = []
for i in range(0, 4):
#    hModis.append(pylab.plot(modisWaveLen[i], modisRsr[i], color=colors[i], linestyle='-'))
    hModis.append(pylab.plot(modisWaveLen[i], modisRsr[i], color='k', linestyle='-'))
    pylab.hold('on')
    mask = dmcRsr[:, i] > 0.001
#    hDmc.append(pylab.plot(dmcWaveLen[mask], dmcRsr[mask, i], color=colors[i], linestyle='--'))
    hDmc.append(pylab.plot(dmcWaveLen[mask], dmcRsr[mask, i], color='k', linestyle='--'))
pylab.legend((hModis[0][0], hDmc[0][0]), ('MODIS', 'DMC'), fontsize=fontSize-2.)
pylab.xlabel('Wavelength ($\mu m$)')
pylab.ylabel('Relative spectral response')
pylab.tight_layout()

f1.savefig('C:/Data/Development/Projects/PhD GeoInformatics/Docs/My Docs/Thesis/Retrieval of Surface Reflectance '
           'from Aerial Imagery/Figure 2 - MODIS and DMC RSRs.eps', dpi=600)


pylab.figure('MODIS, DMC and SPOT Spectral Sensitivities')
colors = ['k','r','g','b']
hModis = []
hDmc = []
hSpot = []
for i in range(0, 4):
#    hModis.append(pylab.plot(modisWaveLen[i], modisRsr[i], color=colors[i], linestyle='-'))
    hModis.append(pylab.plot(modisWaveLen[i], modisRsr[i], color='k', linestyle='-'))
    pylab.hold('on')
    mask = dmcRsr[:, i] > 0.001
#    hDmc.append(pylab.plot(dmcWaveLen[mask], dmcRsr[mask, i], color=colors[i], linestyle='--'))
    hDmc.append(pylab.plot(dmcWaveLen[mask], dmcRsr[mask, i], color='k', linestyle='--'))
    if i<np.shape(spotRsr)[1]:
        mask = np.array(spotRsr)[:, i] > 0.001
        hSpot.append(pylab.plot(np.array(spotWaveLen)[mask], np.array(spotRsr)[mask, i], color=colors[i], linestyle=':'))

pylab.legend((hModis[0][0], hDmc[0][0], hSpot[0][0]), ('MODIS', 'DMC', 'SPOT5'))
pylab.xlabel('Wavelength ($\mu m$)')
pylab.ylabel('Relative spectral response')
pylab.tight_layout()

pylab.figure('MODIS, DMC and QB Spectral Sensitivities')
colors = ['k','r','g','b']
hModis = []
hDmc = []
hSpot = []
for i in range(0, 4):
#    hModis.append(pylab.plot(modisWaveLen[i], modisRsr[i], color=colors[i], linestyle='-'))
    hModis.append(pylab.plot(modisWaveLen[i], modisRsr[i], color='k', linestyle='-'))
    pylab.hold('on')
    mask = dmcRsr[:, i] > 0.001
#    hDmc.append(pylab.plot(dmcWaveLen[mask], dmcRsr[mask, i], color=colors[i], linestyle='--'))
    hDmc.append(pylab.plot(dmcWaveLen[mask], dmcRsr[mask, i], color='k', linestyle='--'))
    if i<np.shape(qbRsr)[1]:
        mask = np.array(qbRsr)[:, i] > 0.001
        hSpot.append(pylab.plot(np.array(qbWaveLen)[mask], np.array(qbRsr)[mask, i], color=colors[i], linestyle=':'))

pylab.legend((hModis[0][0], hDmc[0][0], hSpot[0][0]), ('MODIS', 'DMC', 'QB'))
pylab.xlabel('Wavelength ($\mu m$)')
pylab.ylabel('Relative spectral response')
pylab.tight_layout()

f1 = pylab.figure('DMC and SPOT Spectral Sensitivities')
f1.set_size_inches(6, 4.5, forward=True)
colors = ['k','r','g','b']
# hModis = []
hDmc = []
hSpot = []
for i in range(0, 4):
    # hModis.append(pylab.plot(modisWaveLen[i], modisRsr[i], color=colors[i], linestyle='-'))
    mask = dmcRsr[:, i] > 0.001
    # hDmc.append(pylab.plot(dmcWaveLen[mask], dmcRsr[mask, i], color=colors[i], linestyle='--'))
    hDmc.append(pylab.plot(dmcWaveLen[mask], dmcRsr[mask, i], color='k', linestyle='--'))
    pylab.hold('on')
    if i<np.shape(spotRsr)[1]:
        mask = np.array(spotRsr)[:, i] > 0.001
        # hSpot.append(pylab.plot(np.array(spotWaveLen)[mask], np.array(spotRsr)[mask, i], color=colors[i], linestyle='-'))
        hSpot.append(pylab.plot(np.array(spotWaveLen)[mask], np.array(spotRsr)[mask, i], color='k', linestyle='-'))

pylab.legend((hDmc[0][0], hSpot[0][0]), ('DMC', 'SPOT5'), fontsize=fontSize-2.)
pylab.xlabel('Wavelength ($\mu m$)')
pylab.ylabel('Relative spectral response')
pylab.tight_layout()

f1.savefig('C:/Data/Development/Projects/PhD GeoInformatics/Docs/My Docs/Thesis/Retrieval of Surface Reflectance '
           'from Aerial Imagery/Figure 11 - DMC and SPOT5 RSRs.eps', dpi=600)


pylab.figure('MODIS, DMC and Landsat Spectral Sensitivities')
colors = ['k','r','g','b']
hModis = []
hDmc = []
hSpot = []
for i in range(0, 4):
#    hModis.append(pylab.plot(modisWaveLen[i], modisRsr[i], color=colors[i], linestyle='-'))
    hModis.append(pylab.plot(modisWaveLen[i], modisRsr[i], color='k', linestyle='-'))
    pylab.hold('on')
    mask = dmcRsr[:, i] > 0.001
#    hDmc.append(pylab.plot(dmcWaveLen[mask], dmcRsr[mask, i], color=colors[i], linestyle='--'))
    hDmc.append(pylab.plot(dmcWaveLen[mask], dmcRsr[mask, i], color='k', linestyle='--'))
    if i<np.shape(landsatRsrList)[0]:
        mask = np.array(landsatRsrList[i]) > 0.001
        hSpot.append(pylab.plot(np.array(landsatWaveLenList[i])[mask], np.array(landsatRsrList[i])[mask], color=colors[i], linestyle=':'))

pylab.legend((hModis[0][0], hDmc[0][0], hSpot[0][0]), ('MODIS', 'DMC', 'Landsat'))
pylab.xlabel('Wavelength ($\mu m$)')
pylab.ylabel('Relative spectral response')
pylab.tight_layout()


# MODIS, DMC RSR and ASTER spectra for paper
# pylab.close('all')
f1 = pylab.figure('MODIS and DMC Spectral Sensitivities')
colors = ['orange','r','g','b']
hModis = []
hDmc = []
hAster = []
textWavelen = 910.

ax1 = pylab.axes()
ax2 = ax1.twinx()
for i in range(0, 4):
    hModis.append(ax1.plot(modisWaveLen[i], modisRsr[i], color=colors[i], linestyle='-', axes=ax1))
    pylab.hold('on')
    mask = dmcRsr[:, i] > 0.001
    dmcRsrNorm = dmcRsr[mask, i]/dmcRsr[mask, i].max()
    hDmc.append(ax1.plot(dmcWaveLen[mask], dmcRsrNorm, color=colors[i], linestyle='--', axes=ax1))

hModis.append(ax1.plot(0, 0, color='k', linestyle='-', axes=ax1))  # dummy for legend
hDmc.append(ax1.plot(0, 0, color='k', linestyle='--', axes=ax1))  # dummy for legend

for i in range(0, asterWaveLen.__len__()):
    awvl = asterWaveLen[i]*1000.
    asterMask = (awvl >= dmcWaveLen.min()) & (awvl <= dmcWaveLen.max())
    hAster.append(ax2.plot(awvl[asterMask], asterRefl[i][asterMask], color='k', linestyle=':', axes=ax2))
    # place text
    idx = np.argmin(np.abs(awvl-textWavelen))
    ax2.text(textWavelen, asterRefl[i][idx] + 1., asterFn[i], fontsize=fontSize-2.)

ax1.set_ylim(0., 1.)
ax2.set_ylim(0., 100.)
ax1.set_xlim(400., 1020.)
ax2.set_xlim(400., 1020.)
ax1.legend((hModis[-1][0], hDmc[-1][0], hAster[0][0]), ('MODIS RSR', 'DMC RSR', 'Refl. Eg.'), fontsize = 'small',
           loc='upper right')
# ax2.legend(hModis[0:4], ('NIR', 'Red', 'Green', 'Blue'), fontsize = 'small',
#            loc='center right')
ax1.set_xlabel('Wavelength (nm)')
ax1.set_ylabel('Relative spectral response (RSR)')
ax2.set_ylabel('Reflectance (%)')
mpl.rcParams.update({'font.size': fontSize})
pylab.tight_layout()

# MODIS, DMC, SPOT RSR and ASTER spectra for paper
pylab.close('all')
f1 = pylab.figure('MODIS, DMC, SPOT Spectral Sensitivities and Surf Refl Eg')
colors = ['orange','r','g','b']
hModis = []
hDmc = []
hSpot = []
hAster = []
textWavelen = 910.
fontSize = 14.
ax1 = pylab.axes()
ax2 = ax1.twinx()
for i in range(0, 4):
    hModis.append(ax1.plot(modisWaveLen[i], modisRsr[i], color=colors[i], linestyle='-', marker='', axes=ax1))
    pylab.hold('on')
    mask = dmcRsr[:, i] > 0.001
    dmcRsrNorm = dmcRsr[mask, i]/dmcRsr[mask, i].max()
    hDmc.append(ax1.plot(dmcWaveLen[mask], dmcRsrNorm, color=colors[i], linestyle='--', marker='', axes=ax1))
    if i<np.shape(spotRsr)[1]:
        mask = np.array(spotRsr)[:, i] > 0.001
        hSpot.append(ax1.plot(np.array(spotWaveLen)[mask], np.array(spotRsr)[mask, i], color=colors[i],
                                linestyle=':', marker='', axes=ax1))

hModis.append(ax1.plot(0, 0, color='k', linestyle='-', marker='', axes=ax1))  # dummy for legend
hDmc.append(ax1.plot(0, 0, color='k', linestyle='--', marker='', axes=ax1))  # dummy for legend
hSpot.append(ax1.plot(0, 0, color='k', linestyle=':', marker='', axes=ax1))  # dummy for legend

if True:
    for i in range(0, asterWaveLen.__len__()):
        awvl = asterWaveLen[i]*1000.
        asterMask = (awvl >= dmcWaveLen.min()) & (awvl <= dmcWaveLen.max())
        hAster.append(ax2.plot(awvl[asterMask], asterRefl[i][asterMask], color='k', linestyle='-.', axes=ax2))
        # place text
        idx = np.argmin(np.abs(awvl-textWavelen))
        ax2.text(textWavelen, asterRefl[i][idx] + 1., asterFn[i], fontsize=fontSize-2.)
else:
    for i in range(0, asterWaveLenF.__len__()):
        awvl = asterWaveLenF[i]*1000.
        asterMask = (awvl >= dmcWaveLen.min()) & (awvl <= dmcWaveLen.max())
        hAster.append(ax2.plot(awvl[asterMask], asterReflF[i][asterMask], color='k', linestyle='-.', axes=ax2))
        # place text
        idx = np.argmin(np.abs(awvl-textWavelen))
        # ax2.text(textWavelen, asterReflF[i][idx] + 1., asterFnF[i], fontsize=fontSize-2.)
        ax2.text(textWavelen, asterReflF[i][idx] + 1., asterFnF[i], fontsize=fontSize-2.)

ax1.set_ylim(0., 1.)
ax2.set_ylim(0., 100.)
ax1.set_xlim(400., 1020.)
ax2.set_xlim(400., 1020.)
ax1.legend((hModis[-1][0], hDmc[-1][0], hSpot[-1][0], hAster[0][0]), ('MODIS RSR', 'DMC RSR', 'SPOT RSR', 'Refl. Eg.'), fontsize = 'small',
           loc='upper right')
# ax2.legend(hModis[0:4], ('NIR', 'Red', 'Green', 'Blue'), fontsize = 'small',
#            loc='center right')
ax1.set_xlabel('Wavelength (nm)')
ax1.set_ylabel('Relative spectral response (RSR)')
ax2.set_ylabel('Reflectance (%)')
mpl.rcParams.update({'font.size': fontSize})
pylab.tight_layout()