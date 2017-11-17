
deadwoodFileName = "C:/Data/Development/Projects/PhD GeoInformatics/Data/GEF Field Trial/GEF_Deadwood_2017.10.16_Mdoda.xlsx"

from openpyxl import load_workbook
import numpy as np
import pylab
from scipy.stats import gaussian_kde
import collections
from csv import DictWriter
from collections import OrderedDict


def EvalRecordCs(record, plotArea=(10. ** 2) / (100. ** 2)  ):   # plotArea in ha
    # formula from CDM doc: A/R Methodological Tool - Eq 6
    # equation uses various units!!
    # thumb suck params

    cfTree = 0.48  # carbon fraction tree biomass
    Dj = 0.5   # basic wood density (guess from a quick look at literature) (t d.m/m^3)
                # Note that basic density is similar to the wet/dry ration in Marius' table
    L = 100.   # length of transects (m)
    densityReduction = [1., 0.8, 0.45]
    yc = plotArea * (44./12) * cfTree * Dj * ((np.pi**2)/(8.*L)) * ((record['diameter']/10.)**2) * densityReduction[record['densityState']-1]
    return yc


def EvalPlotCs(allometricModels, plot):
    plotYc = []
    for record in plot:
        yc = EvalRecordCs(allometricModels, record)
        if yc > 0.:
            plotYc.append(yc)
    return plotYc


fontSize = 16

wb = load_workbook(deadwoodFileName)

sheetNames = ['INT-Lying Deadwood', 'INT-Standing Deadwood', 'TCH-Lying Deadwood', 'TCH-Standing Deadwood']
plots = collections.OrderedDict()
for sheetName in sheetNames:
    ws = wb[sheetName]
    for colStart in np.arange(0, ws.max_column, 2):
        if ws[7][colStart].value is None:
            break
        plotName = str(ws[7][colStart].value)
        plot = []
        print sheetName, ' - ', plotName
        for r in ws[9:ws.max_row]:
            if r[colStart].value is None:
                break
            record = OrderedDict()
            record['diameter'] = r[colStart].value
            record['densityState'] = r[colStart + 1].value
            if record['densityState'] == 'D':
                record['densityState'] = 3L
            elif record['densityState'] == 'C':
                record['densityState'] = 2L
            elif record['densityState'] == 'B':
                record['densityState'] = 1L
            elif record['densityState'] == 'A':
                record['densityState'] = 1L
            elif type(record['densityState']) is not long:
                print 'WARNING: unknown density state: ', record['densityState']

            if plotName.startswith('INT'):
                record['yc'] = EvalRecordCs(record, plotArea=(10. ** 2) / (100. ** 2))
            else:
                record['yc'] = EvalRecordCs(record, plotArea=(10. ** 2) / (100. ** 2))

            plot.append(record)
            print '.',
        if plots.__contains__(plotName):
            plots[plotName] = plots[plotName] + plot  # concat lying and standing deadwood
        else:
            plots[plotName] = plot
        print ''
wb = None

# make csv files of results
for plotName in plots.keys():
    plot = plots[plotName]
    outFileName = 'C:\Data\Development\Projects\PhD GeoInformatics\Code\Results\Baviaans2017FieldTrialAnalysis\%s - Deadwood.csv' % (plotName)
    with open(outFileName, 'wb') as outfile:
        writer = DictWriter(outfile, plot[0].keys())
        writer.writeheader()
        writer.writerows(plot)

# vars = [model['vars'] for model in allometricModels.values()]
# print np.unique(vars)
pylab.close('all')
i = 1
ycTtl = 0.
measTtl = 0

pylab.figure()
for plotKey, plot in plots.iteritems():
    height = np.float64([record['diameter'] for record in plot])
    yc = np.float64([record['yc'] for record in plot])
    ycTtl += yc.sum()
    measTtl += yc.__len__()
    kde = gaussian_kde(height)  #, bw_method=bandwidth / height.std(ddof=1))
    heightGrid = np.linspace(0, 500, 100)
    heightKde = kde.evaluate(heightGrid)
    pylab.subplot(2, 3, i)
    pylab.plot(heightGrid, heightKde)
    axLim = pylab.axis()
    # h = pylab.plot([50, 50], [0, heightKde.max()], 'r')
    pylab.grid('on')
    pylab.xlabel('Dead Wood Diam. (mm)', fontdict={'size':fontSize})
    pylab.ylabel('Prob.(diam.)', fontdict={'size':fontSize})
    pylab.title('Diameter distribution for plot %s' % (plotKey), fontdict={'size':fontSize})
    # if i >= plots.__len__():
    #     pylab.legend(h, ['50mm threshold'], loc='upper left', bbox_to_anchor=(1.2, 1), prop={'size':fontSize})
    pylab.axis([axLim[0], axLim[1], 0, heightKde.max()])
    i += 1

cutOffTtlYc = 0.
cutOffTtlMeas = 0
i = 1
pylab.figure()
for plotKey, plot in plots.iteritems():
    yc = np.array([record['yc'] for record in plot])
    height = np.float64([record['diameter'] for record in plot])
    idx = np.argsort(height)

    ycCumSum = np.cumsum(yc[idx])
    pylab.subplot(2, 3, i)
    pylab.plot(height[idx], ycCumSum)
    axLim = pylab.axis()
    # h = pylab.plot([50, 50], [axLim[2], axLim[3]], 'r')
    pylab.axis(axLim)
    pylab.grid('on')
    pylab.xlabel('Dead Wood Diam. (mm)', fontdict={'size':fontSize})
    pylab.ylabel('Cum. Distr.(C. stock) (kg)', fontdict={'size':fontSize})
    # pylab.xlim([0, 350])
    pylab.title('Diameter / C. stock relation for plot %s' % (plotKey), fontdict={'size':fontSize})
    # if i >= plots.__len__():
    #     pylab.legend(h, ['50cm threshold'], loc='upper left', bbox_to_anchor=(1.2, 1), prop={'size':fontSize})
    i += 1

    idx = height > 20
    cutOffTtlYc += yc[idx].sum()
    cutOffTtlMeas += (~idx).sum()
    print str(plotKey), ": "
    print "Ttl: ", str(yc.sum())
    print "Hgt>20: ", str(yc[idx].sum())
    print "Hgt>20/Plot Ttl: ", str(yc[idx].sum()/yc.sum())
    print "Hgt>20/Ttl: ", str(yc[idx].sum()/ycTtl)

print "---------------------------------------------"
print "Ttl C Hgt>20/Ttl: ", str(100.*cutOffTtlYc/ycTtl)
print "Ttl Meas Hgt>20/Ttl: ", str(100.*float(cutOffTtlMeas)/float(measTtl))

