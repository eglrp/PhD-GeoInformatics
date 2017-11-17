# ----------------------------------------------------------------------------------------------------------------------
# Read in Allometric model parameters

allometryFileName = "C:/Data/Development/Projects/PhD GeoInformatics/Data/GEF Field Trial/AllometricModels.xlsx"
woodyFileName = "C:/Data/Development/Projects/PhD GeoInformatics/Data/GEF Field Trial/GEF_Woody canopy_2017.10.16_Mdoda.xlsx"

from openpyxl import load_workbook
import numpy as np
import pylab
from scipy.stats import gaussian_kde
import collections
from csv import DictWriter
from collections import OrderedDict
fontSize = 16

def EvalRecordCs(allometricModels, record):
    # vars = [model['vars'] for model in allometricModels.values()]
    if not allometricModels.__contains__(record['species']):
        print record['species'], " not found"
        return 0.
    model = allometricModels[record['species']]
    x = 0.
    CD = np.mean([record['canopyLength'], record['canopyWidth']])
    CA = np.pi*(CD/2)**2
    if model['vars'] == 'CA.H':
        x = CA*record['height']
    elif model['vars'] == 'CA.SL':
        x = CA*record['height']
    elif model['vars'] == 'CD':
        x = CD
    elif model['vars'] == 'CD.H':
        x = CD*record['height']
    elif model['vars'] == 'Hgt':
        x = record['height']
    else:
        print model['vars'], " unknown variable"
        return 0.
    yn = np.exp(model['ay'])*x**model['by']   # "naive"
    Yc = yn*model['MB']

    if model['useWdRatio']:
        if model.__contains__('wdRatio'):
            Yc = Yc * model['wdRatio']
        else:
            print "WD Ratio for ", record['species'], " not found - using 1."

    return Yc

def EvalPlotCs(allometricModels, plot):
    plotYc = []
    for record in plot:
        Yc = EvalRecordCs(allometricModels, record)
        if Yc > 0.:
            plotYc.append(Yc)
        print record['species'], " - ", str(Yc)
    return plotYc

# ---------------------------------------------------------------------------------------------------------------------


wb = load_workbook(allometryFileName)
ws = wb["Allometric Models"]

first_row = ws[0 + 1]
header = []
for c in first_row:
    header.append(c.value)

allometricModels = {}
for r in ws[2:ws.max_row]:  #how to find num rows?
    if r[0].value is None:
        break
    species = str.strip(str(r[0].value[0]) + '. ' + str(r[1].value))
    model = {}
    model['vars'] = r[3].value
    model['ay'] = r[6].value
    model['by'] = r[7].value
    model['MB'] = r[13].value
    model['duan'] = r[12].value
    if str(r[14].value) == 'x':
        model['useWdRatio'] = False
    else:
        model['useWdRatio'] = True
    allometricModels[species] = model
ws = None

ws = wb["Wet Dry Ratios"]
first_row = ws[0 + 1]
header = []
for c in first_row:
    header.append(c.value)

wdRatios = {}
for r in ws[2:ws.max_row]:  #how to find num rows?
    if r[0].value is None:
        break
    species = str.strip(str(r[0].value[0]) + '. ' + str(r[1].value))
    model = {}
    model['WDratio'] = r[4].value
    if allometricModels.__contains__(species):
        allometricModels[species]['wdRatio'] = r[4].value
    else:
        print species, ' not found in allometric models'
    wdRatios[species] = model
ws = None

wb = None

# ----------------------------------------------------------------------------------------------------------------------
# Read in trial woody measurements

wb = load_workbook(woodyFileName)

plots = collections.OrderedDict()
for ws in wb:
    print ws.title, ' rows: ', ws.max_row
    first_row = ws[0 + 5]
    header = []
    for c in first_row:
        header.append(c.value)

    plot = []
    for r in ws[6:ws.max_row]:
        if r[1].value is None:
            break
        record = OrderedDict()
        species = str(r[1].value).strip()
        # hack for unknown plants
        if species == 'A. ferox' or species.__contains__('Aloe'):
            species = 'A. speciosa'
        # if species.__contains__('unk'):
        if species.__contains__('Asparagus') or species.__contains__('hairy'):
            species = 'A. capensis'
        if species.__contains__('Crassula') or species.__contains__('horns'):
            species = 'C. ovata'

        record['species'] = str(species).strip()
        record['canopyWidth'] = r[2].value
        record['canopyLength'] = r[3].value
        record['height'] = r[4].value
        record['bsd'] = r[5].value
        yc = EvalRecordCs(allometricModels, record)
        record['yc'] = yc
        plot.append(record)
    plots[ws.title] = plot
    outFileName = 'C:\Data\Development\Projects\PhD GeoInformatics\Code\Results\Baviaans2017FieldTrialAnalysis\%s - Woody.csv' % (ws.title)
    with open(outFileName,'wb') as outfile:
        writer = DictWriter(outfile, plot[0].keys())
        writer.writeheader()
        writer.writerows(plot)

wb = None

# vars = [model['vars'] for model in allometricModels.values()]
# print np.unique(vars)
pylab.close('all')
i = 1
ycTtl = 0.
pylab.figure()
for plotKey, plot in plots.iteritems():
    yc = np.array([record['yc'] for record in plot])
    height = np.float64([record['height'] for record in plot])
    ycTtl += yc.sum()

    kde = gaussian_kde(height)  #, bw_method=bandwidth / height.std(ddof=1))
    heightGrid = np.linspace(0, 300, 100)
    heightKde = kde.evaluate(heightGrid)
    pylab.subplot(2, 3, i)
    pylab.plot(heightGrid, heightKde)
    axLim = pylab.axis()
    h = pylab.plot([50, 50], [0, heightKde.max()], 'r')
    pylab.grid('on')
    pylab.xlabel('Plant Height (cm)', fontdict={'size':fontSize})
    pylab.ylabel('Prob.(height)', fontdict={'size':fontSize})
    pylab.title('Height distribution for plot %s' % (plotKey), fontdict={'size':fontSize})
    if i >= plots.__len__():
        pylab.legend(h, ['50cm threshold'], loc='upper left', bbox_to_anchor=(1.2, 1), prop={'size':fontSize})
    pylab.axis([axLim[0], axLim[1], 0, heightKde.max()])
    i += 1

i = 1
pylab.figure()
for plotKey, plot in plots.iteritems():
    yc = np.array([record['yc'] for record in plot])
    height = np.float64([record['height'] for record in plot])
    idx = np.argsort(height)

    ycCumSum = np.cumsum(yc[idx])
    pylab.subplot(2, 3, i)
    pylab.plot(height[idx], ycCumSum)
    axLim = pylab.axis()
    h = pylab.plot([50, 50], [axLim[2], axLim[3]], 'r')
    pylab.axis(axLim)
    pylab.grid('on')
    pylab.xlabel('Plant Height (cm)', fontdict={'size':fontSize})
    pylab.ylabel('Cum. Distr.(C. stock) (kg)', fontdict={'size':fontSize})
    pylab.xlim([0, 350])
    pylab.title('Height / C. stock relation for plot %s' % (plotKey), fontdict={'size':fontSize})
    if i >= plots.__len__():
        pylab.legend(h, ['50cm threshold'], loc='upper left', bbox_to_anchor=(1.2, 1), prop={'size':fontSize})
    i += 1

    idx = height > 50
    print str(plotKey), ": "
    print "Ttl: ", str(yc.sum())
    print "Hgt>50: ", str(yc[idx].sum())
    print "Hgt>50/Plot Ttl: ", str(yc[idx].sum()/yc.sum())
    print "Hgt>50/Ttl: ", str(yc[idx].sum()/ycTtl)

#----------------------------------------------------------------------------------------------------------------------
# Analyse edge behaviour

# find the area of a slice of the ellipse (x/a)**2 + (x/b)**2 = 1 where the slice is theta = (0 to theta)
def EllipseOriginSliceArea(a, b, theta):
    #area = (0.5) * a * b * (0.5*np.sin(2*theta) + theta)
    area = a*b*0.5*(theta - np.arctan((b-a)*np.sin(2*theta)/(b+a+(b-a)*np.cos(2*theta))))
    return area

# find the area of a slice of the ellipse (x/a)**2 + (x/b)**2 = 1 where the slice is theta = (thetaStart to thetaEnd)
def EllipseSliceArea(a, b, thetaStart, thetaEnd):
    return EllipseOriginSliceArea(a, b, thetaEnd) - EllipseOriginSliceArea(a, b, thetaStart)

# find the area of a section of the ellipse (x/a)**2 + (x/b)**2 = 1 where the section is x = (-a to x_section)
def EllipseSectionArea(a, b, xSection):
    xSign = 1.
    if xSection < 0:
        print "WARNING: xSection should positive"
        xSection = np.abs(xSection)
        xSign = -1.
    if xSection > a:
        print "WARNING: xSection should be less than a"
        xSection = a

    ySection = b*np.sqrt(1-(xSection/a)**2)
    triArea = 0.5*xSection*ySection
    ellipseSliceArea = EllipseSliceArea(a, b, np.arctan(ySection/xSection), np.pi)
    ellipseSectionArea = 2*(triArea + ellipseSliceArea)
    ellipseArea = np.pi*a*b

    ellipseOtherSliceArea = EllipseSliceArea(a, b, 0, np.arctan(ySection/xSection))
    ellipseOtherSectionArea = 2*(ellipseOtherSliceArea - triArea)
    print ellipseOtherSectionArea + ellipseSectionArea
    print ellipseArea
    # sectionLength = 2*ySection
    print ellipseSectionArea, ySection
    if xSign < 0:
        ellipseSectionArea = ellipseOtherSectionArea

    return ellipseSectionArea, ySection

def CircleSectionArea(a, b, xSection):
    a = b = (a + b)/2   # Simply for Marius' formula for CD/2
    xSign = 1.
    if xSection < 0:
        print "WARNING: xSection should positive"
        xSection = np.abs(xSection)
        xSign = -1.
    if xSection > a:
        print "WARNING: xSection should be less than a"
        xSection = a

    ySection = b*np.sqrt(1 - (xSection/a)**2)
    triArea = 0.5*xSection*ySection
    ellipseSliceArea = EllipseSliceArea(a, b, np.arctan(ySection/xSection), np.pi)
    ellipseSectionArea = 2*(triArea + ellipseSliceArea)
    ellipseArea = np.pi*a*b

    ellipseOtherSliceArea = EllipseSliceArea(a, b, 0, np.arctan(ySection/xSection))
    ellipseOtherSectionArea = 2*(ellipseOtherSliceArea - triArea)
    print ellipseOtherSectionArea + ellipseSectionArea
    print ellipseArea
    if xSign < 0:
        ellipseSectionArea = ellipseOtherSectionArea
    # sectionLength = 2*ySection
    return ellipseSectionArea, ySection

#---------------------------------------------------------------------------------------------------------------
# simulate ellipse section areas and areas of full ellipses with a = (a+xSection)/2 and b=ySection
# to see how close the areas are
from scipy import stats
av = np.abs(np.random.randn(100)+0.5)
bv = np.abs(np.random.randn(100)+0.5)
xSectionV = 0.5*(av+bv)*np.abs(0.3*np.random.randn(100))
all(xSectionV<av)

sectionAreaV = []
circleAreaV = []
for a, b, xSection in zip(av, bv, xSectionV):
    sa, bnew = CircleSectionArea(a, b, xSection)
    sectionAreaV.append(sa)
    # circleAreaV.append((a+xSection)*0.5*bnew*4)  # rectangular area
    circleAreaV.append(np.pi*(((a+xSection)*0.5 + bnew)/2)**2)

sectionAreaV = np.array(sectionAreaV)
circleAreaV = np.array(circleAreaV)

(slope, intercept, r, p, stde) = stats.linregress(sectionAreaV, circleAreaV)

pylab.figure()
# pylab.subplot(1, 2, 1)
# pylab.plot(av, bv, 'kx')
# pylab.subplot(1, 2, 2)
pylab.plot(sectionAreaV, circleAreaV, 'bx')
m = np.max([sectionAreaV.max(), circleAreaV.max()])
h, = pylab.plot([0, m], [0, m], 'r')
pylab.text(m*0.6, m*0.1, str.format('$R^2$ = {0:.2f}', np.round(r ** 2, 2)), fontdict={'size':fontSize})
pylab.grid()
pylab.legend([h], ['1:1'], prop={'size':fontSize})
pylab.xlabel('Canopy Area to Edge', fontdict={'size':fontSize})
pylab.ylabel('Approx. Circular Area', fontdict={'size':fontSize})
pylab.title('Edge intersected canopy area approximation', fontdict={'size':fontSize})
# for model, species in allometricModels.iteritems():


#---------------------------------------------------------------------------------------------------------------
# simulate how linearly allometric models behave to scaling in canopy area / diameter

# first take a look at how much each species contributes

allSpecies = np.array([k for k, v in allometricModels.iteritems()])
allSpeciesYc = dict(zip(allSpecies, np.zeros(allSpecies.shape)))
allSpecies = np.array([k for k, v in allSpeciesYc.iteritems()])  # order can change!!!

for plotKey, plot in plots.iteritems():
    for record in plot:
        if allSpecies.__contains__(record['species']):
            if record['species'].__contains__('junceum'):
                print 'JUNCEUM'
                break
            allSpeciesYc[record['species']] += record['yc']


pylab.figure()
pylab.bar(range(0, allSpecies.size), [v for k,v in allSpeciesYc.iteritems()])
# pylab.yscale('log')
pylab.xticks(range(0, allSpecies.size), allSpecies, rotation='vertical', fontSize=fontSize)  #prop={'size':fontSize-2})
pylab.grid('on')
pylab.ylabel('Total C (kg)', fontdict={'size':fontSize})
pylab.title('Total Trial C per Species', fontdict={'size':fontSize})

#find 4 highest contributing species
idx = np.flipud(np.argsort([v for k,v in allSpeciesYc.iteritems()]))
topSpecies = allSpecies[idx[:4]]

# -------------------------------------------------------------------------------------------------------
# see how y varies with x for these species


# a hacked version of the fn above to return y for varying x
def EvalRecordCsV(allometricModels, record):
    # vars = [model['vars'] for model in allometricModels.values()]
    if not allometricModels.__contains__(record['species']):
        print record['species'], " not found"
        return 0.
    model = allometricModels[record['species']]
    x = 0.
    CD = np.mean([record['canopyLength'], record['canopyWidth']])  #make these the max valies
    CD = np.linspace(10, CD, 100)
    CA = np.pi*(CD/2)**2
    if model['vars'] == 'CA.H':
        x = CA*record['height']
    elif model['vars'] == 'CA.SL':
        x = CA*record['height']
    elif model['vars'] == 'CD':
        x = CD
    elif model['vars'] == 'CD.H':
        x = CD*record['height']
    elif model['vars'] == 'Hgt':
        x = record['height']
    else:
        print model['vars'], " unknown variable"
        return 0.
    yn = np.exp(model['ay'])*x**model['by']   # "naive"
    yc = yn*model['MB']
    if model['useWdRatio']:
        if model.__contains__('wdRatio'):
            yc = yc * model['wdRatio']
        else:
            print "WD Ratio for ", record['species'], " not found - using 1."

    return yc, x


pylab.figure()
pi = 0
for specie in topSpecies:
    model = allometricModels[specie]
    # max values
    record['canopyLength'] = 200
    record['canopyWidth'] = 200
    record['height'] = 150
    record['species'] = specie

    yc,x = EvalRecordCsV(allometricModels, record)
    pylab.subplot(2, 2, pi+1)
    pylab.plot(x, yc)
    pylab.grid()
    pylab.ylabel('Yc (kg)', fontdict={'size':fontSize})
    pylab.xlabel(allometricModels[record['species']]['vars'], fontdict={'size':fontSize})
    pylab.title(specie, fontdict={'size':fontSize})
    pi += 1




#------------------------------------------------------------------------------------------------------------------
# suspect code below
pylab.figure()
pi = 0
for specie in topSpecies:
    model = allometricModels[specie]
    canopyLengths = []
    canopyLengths = np.linspace(20,300,100)
    ycs = []
    for canopyLength in canopyLengths:  # i in range(0, 100):
        record = {}
        if False:
            record['canopyLength'] = np.random.rand(1)*200
            canopyLengths.append(record['canopyLength'])
        else:
            record['canopyLength'] = canopyLength
        record['canopyWidth'] = canopyLength*0.5
        record['height'] = 150
        record['species'] = specie

        edgeSection = record['canopyLength']*0.2
        sectionArea, ySection = CircleSectionArea(edgeSection + 0.5*record['canopyLength'], 0.5*record['canopyWidth'], edgeSection)
        record['canopyWidth'] = edgeSection*2

        # get yc for full canopy
        ycFull = EvalRecordCs(allometricModels, record)   # find c for full canopy

        ycs.append(ycFull)

    ycs = np.array(ycs)
    pylab.subplot(2, 2, pi + 1)
    pylab.plot(canopyLengths, ycs)
    pylab.plot([canopyLengths.min(), canopyLengths.max()], [ycs.min(), ycs.max()])
    pylab.grid()
    pylab.ylabel('Yc')
    pylab.xlabel('Canopy Length')
    pylab.title(specie)
    pi += 1

    if False:

        # update record for section of canopy and get yc
        sectionArea, ySection = CircleSectionArea(record['canopyLength']/2, record['canopyWidth']/2, edgeSection)
        record['canopyLength'] = record['canopyLength'] - edgeSection
        record['canopyWidth'] = ySection
        ycSectionApprox = EvalRecordCs(allometricModels, record)   #find c for approx inside area of canopy

        # get yc for actual portion of canopy
        model = allometricModels[record['species']]
        x = 0.
        CD = np.mean([record['canopyLength'], record['canopyWidth']])
        CA = sectionArea  #np.pi*(CD/2)**2
        if model['vars'] == 'CA.H':
            x = CA*record['height']
        elif model['vars'] == 'CA.SL':
            x = CA*record['height']
        elif model['vars'] == 'CD':
            x = CD
        elif model['vars'] == 'CD.H':
            x = CD**record['height']
        elif model['vars'] == 'Hgt':
            x = record['height']
        else:
            print model['vars'], " unknown variable"
        yn = model['ay']*x**model['by']   # "naive"
        Yc = yn*model['duan']
