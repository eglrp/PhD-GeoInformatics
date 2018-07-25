# ----------------------------------------------------------------------------------------------------------------------
# this is now a generic script to process any sampling data
# includes woody and litter where available
# ----------------------------------------------------------------------------------------------------------------------
# Read in Allometric model parameters

allometryFileName = "C:/Data/Development/Projects/PhD GeoInformatics/Data/GEF Field Trial/AllometricModels.xlsx"
# woodyFileName = "C:/Data/Development/Projects/PhD GeoInformatics/Data/GEF Sampling/Sampling March 2018/GEF_Woody spp_INTACT_2018.04.24_Mdoda.xlsx"
# woodyFileName = "C:/Data/Development/Projects/PhD GeoInformatics/Data/GEF Sampling/Sampling March 2018/GEF_Woody spp_INTACT_2018.04.24_Mdoda.xlsx"

# woodyFileName = "C:/Data/Development/Projects/PhD GeoInformatics/Data/GEF Sampling/Sampling June 2018/GEF_MV_Woody spp_2018.07.06.Mdoda.xlsx"
# litterFileName = "C:/Data/Development/Projects/PhD GeoInformatics/Data/GEF Sampling/Sampling June 2018/GEF_Litter_MV plots_2018.07.23_Cos.xlsx"
woodyFileName = "C:/Data/Development/Projects/PhD GeoInformatics/Data/GEF Sampling/Sampling June 2018/GEF_SS_Woody spp_2018.07.13.Mdoda.xlsx"
litterFileName = "C:/Data/Development/Projects/PhD GeoInformatics/Data/GEF Sampling/Sampling June 2018/GEF_Litter_SS plots_2018.07.23_Cos.xlsx"

woodyFileName =  "C:/Data/Development/Projects/PhD GeoInformatics/Data/GEF Sampling/Sampling Dec 2017/GEF_Woody spp_2018.03.10_MdodaQC.xlsx"
litterFileName = "C:/Data/Development/Projects/PhD GeoInformatics/Data/GEF Sampling/Sampling Dec 2017/GEF_Litter_2018.03.09_Cos.xlsx"

from openpyxl import load_workbook
import numpy as np
import pylab
from scipy.stats import gaussian_kde
import collections
from csv import DictWriter
from collections import OrderedDict
import os.path
fontSize = 16


def EvalRecordCs(allometricModels, record):
    # vars = [model['vars'] for model in allometricModels.values()]
    if not allometricModels.__contains__(record['species']):
        print record['species'], " not found"
        return 0.
    model = allometricModels[record['species']]
    x = 0.
    CD = np.mean([record['canopyLength'], record['canopyWidth']])
    CA = np.pi*(CD/2)**2
    if model['vars'] == 'CA.H':
        x = CA*record['height']
    elif model['vars'] == 'CA.SL':
        x = CA*record['height']
    elif model['vars'] == 'CD':
        x = CD
    elif model['vars'] == 'CD.H':
        x = CD*record['height']
    elif model['vars'] == 'Hgt':
        x = record['height']
    else:
        print model['vars'], " unknown variable"
        return 0.
    yn = np.exp(model['ay'])*x**model['by']   # "naive"
    Yc = yn*model['MB']

    if model['useWdRatio']:
        if model.__contains__('wdRatio'):
            Yc = Yc * model['wdRatio']
        # else:
        #     print "WD Ratio for ", record['species'], " not found - using 1."

    return Yc

def EvalPlotCs(allometricModels, plot):
    plotYc = []
    for record in plot:
        Yc = EvalRecordCs(allometricModels, record)
        if Yc > 0.:
            plotYc.append(Yc)
        print record['species'], " - ", str(Yc)
    return plotYc

# ---------------------------------------------------------------------------------------------------------------------
# load allometric models

wb = load_workbook(allometryFileName)
ws = wb["Allometric Models"]

first_row = ws[0 + 1]
header = []
for c in first_row:
    header.append(c.value)

allometricModels = {}
for r in ws[2:ws.max_row]:  #how to find num rows?
    if r[0].value is None:
        break
    species = str.strip(str(r[0].value[0]) + '.' + str(r[1].value))
    model = {}
    model['vars'] = r[3].value
    model['ay'] = r[6].value
    model['by'] = r[7].value
    model['MB'] = r[13].value
    model['duan'] = r[12].value
    if str(r[14].value) == 'x':
        model['useWdRatio'] = False
    else:
        model['useWdRatio'] = True
    allometricModels[species] = model
ws = None

ws = wb["Wet Dry Ratios"]
first_row = ws[0 + 1]
header = []
for c in first_row:
    header.append(c.value)

wdRatios = {}
for r in ws[2:ws.max_row]:  #how to find num rows?
    if r[0].value is None:
        break
    species = str.strip(str(r[0].value[0]) + '.' + str(r[1].value))
    model = {}
    model['WDratio'] = r[4].value
    if allometricModels.__contains__(species):
        allometricModels[species]['wdRatio'] = r[4].value
    else:
        print species, ' not found in allometric models'
    wdRatios[species] = model
ws = None

wb = None

# ----------------------------------------------------------------------------------------------------------------------
# Read in woody measurements

wb = load_workbook(woodyFileName)

# make a list of "not found" species i.e. species without allometric models
notFoundSpecies = {}
for ws in wb:
    print ws.title, ' rows: ', ws.max_row
    first_row = ws[1]
    header = []
    for c in first_row:
        header.append(c.value)

    for r in ws[2:ws.max_row]:
        if r[0].value is None:
            break
        species = str(r[2].value).strip().replace('. ','.')

        if not allometricModels.__contains__(species):
            print species, " not found"
            if notFoundSpecies.has_key(species):
                notFoundSpecies[species] += 1
            else:
                notFoundSpecies[species] = 1

# if MV:
#     outFileName = 'C:/Data/Development/Projects/PhD GeoInformatics/Data/GEF Sampling/Sampling June 2018/Unknown Species MV.csv'
# else:
#     outFileName = 'C:/Data/Development/Projects/PhD GeoInformatics/Data/GEF Sampling/Sampling June 2018/Unknown Species SS.csv'
outFileName = str.format('{0}/{1} - Unknown Species.csv', os.path.dirname(woodyFileName), os.path.splitext(os.path.basename(woodyFileName))[0])

import csv
with open(outFileName, 'wb') as outfile:
    writer = csv.writer(outfile, delimiter=',',
                            quotechar='/',  quoting=csv.QUOTE_MINIMAL)
    writer.writerow(['Species', 'N'])
    for species, N in notFoundSpecies.iteritems():
        writer.writerow([species, N])

print '------------------------------------------------------'


# read in allometry and find woody cs
nestedPlots = {}
for ws in wb:
    print ws.title, ' rows: ', ws.max_row
    first_row = ws[1]
    header = []
    for c in first_row:
        header.append(c.value)

    plots = collections.OrderedDict()
    for r in ws[2:ws.max_row]:
        if r[0].value is None:
            break
        record = OrderedDict()
        species = str(r[2].value).strip().replace('. ','.')

        # hack for unknown plants
        if species == 'A.ferox' or species.__contains__('Aloe') or species.__contains__('A.ferox'):
            species = 'A.speciosa'
        # if species.__contains__('unk'):
        if species.__contains__('Asparagus') or species.__contains__('hairy'):
            species = 'A.capensis'
        if species.__contains__('Crassula') or species.__contains__('horns') or species.__contains__('rupestris'):
            species = 'C.ovata'
        if species == 'Euclea sp1':
            species = 'E.undulata'
        if species.__contains__('Euphorbia') or species.__contains__('Euphorbia'):
            species = 'E.coerulescens'
        if species == 'Schotia sp1':
            species = 'S.afra'
        if str.lower(species).__contains__('grewia'):
            species = 'G.robusta'
        if species.__contains__('Lycium'):
            species = 'L.ferocissimum'
        if species.__contains__('Ysterhout') or species.__contains__('Acacia'):   # very crude guesses
            species = 'P.capensis'
        if species.__contains__('Gymnosporia'):   # very crude guesses
            species = 'G.polyacantha'
        if species.__contains__('Boscia'):
            species = 'B.oleoides'
        if species == 'A.striatus':
            species = 'A.striata'
        if species.__contains__('Rhigozum'):   # very crude guesses
            species = 'R.obovatum'
        if species.__contains__('Searsia') or species == 'S.longspina' or species == 'S.longispina':
            species = 'S.longispina'
        if species.__contains__('Crassula'):
            species = 'C.ovata'

        if not allometricModels.__contains__(species):
            print species, " not found, no subs"


        id = str(r[0].value).strip().replace('-','')
        idNum = np.int32(id[2:])  # get rid of leading zeros
        id = '%s%d' % (id[:2], idNum)

        record['ID'] = id
        record['degrClass'] = str(r[1].value)
        record['species'] = str(species).strip()
        record['canopyWidth'] = r[3].value
        record['canopyLength'] = r[4].value
        record['height'] = r[5].value

        # zerp empty records
        fields = ['canopyWidth', 'canopyLength', 'height']
        for f in fields:
            if record[f] is None:
                record[f] = 0

        if r.__len__() > 6:
            record['bsd'] = str(r[6].value)
        else:
            record['bsd'] = ""
        yc = EvalRecordCs(allometricModels, record)
        record['yc'] = yc
        if plots.has_key(id):
            plots[id].append(record)
        else:
            plots[id] = [record]
    nestedPlots[ws.title] = plots

    # outFileName = 'C:/Data/Development/Projects/PhD GeoInformatics/Code/Results/Baviaans2017FieldTrialAnalysis/%s - Woody.csv' % (ws.title)
    # with open(outFileName,'wb') as outfile:
    #     writer = DictWriter(outfile, plot[0].keys())
    #     writer.writeheader()
    #     writer.writerows(plot)
    # if MV:
    #     outFileName = 'C:/Data/Development/Projects/PhD GeoInformatics/Data/GEF Sampling/Sampling June 2018/%s - MV Woody.csv' % (ws.title)
    # else:
    #     outFileName = 'C:/Data/Development/Projects/PhD GeoInformatics/Data/GEF Sampling/Sampling June 2018/%s - SS Woody.csv' % (ws.title)
    outFileName = str.format('{0}/{1} - Woody.csv', os.path.dirname(woodyFileName),
                             os.path.splitext(os.path.basename(woodyFileName))[0])

    with open(outFileName,'wb') as outfile:
        writer = DictWriter(outfile, plots.values()[0][0].keys())
        writer.writeheader()
        for plot in plots.values():
            writer.writerows(plot)

wb = None
#----------------------------------------------------------------------------------------------------------------
# read in litter data
wb = load_workbook(litterFileName, data_only=True)

litterDict = {}
ws = wb['Litter dry wts']
for r in ws[3:ws.max_row]:
    id = str(r[0].value).strip()
    id = id.replace('-0','')
    id = id.replace('-','')
    if id == '' or id is None or id == 'None':
        print 'No ID, breaking'
        break
    else:
        print id,
    dryWeight = r[12].value
    if litterDict.has_key(id):
        litterDict[id]['dryWeight'] += dryWeight
    else:
        litterDict[id] = {'dryWeight': dryWeight}
wb = None

#---------------------------------------------------------------------------------------------------------------
# extrapolate 5x5m subplots to full size
# (separate contributions from 5x5m plants ><50cm high, extrap <50 to full size, add >50 to full size)
subPlots = nestedPlots['5x5m']
if nestedPlots.has_key('10 x 10m'):
    outerPlots = nestedPlots['10 x 10m']
    outerSize = 10.
else:
    outerPlots = nestedPlots['20 x 20m']
    outerSize = 20.

summaryPlots = {}
for id, subPlot in subPlots.iteritems():
    subHeight = np.array([r['height'] for r in subPlot])
    subYc = np.array([r['yc'] for r in subPlot])
    smallIdx = subHeight < 50
    outerPlot = outerPlots[id]
    # outerHeight = np.array([r['height'] for r in outerPlot])
    outerYc = np.array([r['yc'] for r in outerPlot])
    summaryPlots[id] = {}
    # nb the 4 needs to change for other plot sizes
    # if MV:
    #     outerSize = 10
    # else:
    #     outerSize = 20
    summaryYc = ((outerSize/5.)**2) * subYc[smallIdx].sum() + subYc[np.logical_not(smallIdx)].sum() + outerYc.sum()

    summaryPlots[id]['ID'] = id
    summaryPlots[id]['Yc'] = summaryYc
    summaryPlots[id]['Size'] = outerSize
    summaryPlots[id]['YcHa'] = (100.**2) * summaryYc/(outerSize**2)
    summaryPlots[id]['N'] = ((outerSize/5.)**2) * smallIdx.sum() + (subYc.__len__() - smallIdx.sum()) + outerYc.__len__()
    summaryPlots[id]['LitterDryWeight'] = litterDict[id]['dryWeight']/1000  # g to kg
    summaryPlots[id]['LitterHa'] = summaryPlots[id]['LitterDryWeight'] * (100.**2) / (4 * (0.5**2))
    summaryPlots[id]['AgbHa'] = summaryPlots[id]['LitterHa'] + summaryPlots[id]['YcHa']

# write out summary ground truth for each plot
# if MV:
#     outFileName = 'C:/Data/Development/Projects/PhD GeoInformatics/Data/GEF Sampling/Sampling June 2018/Summary - MV Woody.csv'
# else:
#     outFileName = 'C:/Data/Development/Projects/PhD GeoInformatics/Data/GEF Sampling/Sampling June 2018/Summary - SS Woody.csv'
    outFileName = str.format('{0}/{1} - Summary Woody & Litter.csv', os.path.dirname(woodyFileName),
                             os.path.splitext(os.path.basename(woodyFileName))[0])

with open(outFileName, 'wb') as outfile:
    writer = DictWriter(outfile, summaryPlots.values()[0].keys())
    writer.writeheader()
    writer.writerows(summaryPlots.values())



#-----------------------------------------------------------------------------------------------------------------
#  look at height & yc distribution for 5x5m plots

plots = nestedPlots['5x5m']
# vars = [model['vars'] for model in allometricModels.values()]
# print np.unique(vars)
pylab.close('all')
i = 1
ycTtl = 0.
pylab.figure()
plotSummary = []
for plotKey, plot in plots.iteritems():
    yc = np.array([record['yc'] for record in plot])
    height = np.float64([record['height'] for record in plot])
    ycTtl += yc.sum()
    plotSummary.append({'ID': plotKey.replace('-','_'), 'Yc': yc.sum(), 'N': yc.__len__()})

    kde = gaussian_kde(height)  #, bw_method=bandwidth / height.std(ddof=1))
    heightGrid = np.linspace(0, 300, 100)
    heightKde = kde.evaluate(heightGrid)
    pylab.subplot(4, 5, i)
    pylab.plot(heightGrid, heightKde)
    axLim = pylab.axis()
    h = pylab.plot([50, 50], [0, heightKde.max()], 'r')
    pylab.grid('on')
    pylab.xlabel('Plant Height (cm)', fontdict={'size':fontSize})
    pylab.ylabel('Prob.(height)', fontdict={'size':fontSize})
    pylab.title('Height dist. for %s' % (plotKey), fontdict={'size':fontSize})
    if i >= plots.__len__():
        pylab.legend(h, ['50cm threshold'], loc='upper left', bbox_to_anchor=(1.2, 1), prop={'size':fontSize})
    pylab.axis([axLim[0], axLim[1], 0, heightKde.max()])
    i += 1

i = 1
pylab.figure()
for plotKey, plot in plots.iteritems():
    yc = np.array([record['yc'] for record in plot])
    height = np.float64([record['height'] for record in plot])
    idx = np.argsort(height)

    ycCumSum = np.cumsum(yc[idx])
    pylab.subplot(4, 5, i)
    pylab.plot(height[idx], ycCumSum)
    axLim = pylab.axis()
    h = pylab.plot([50, 50], [axLim[2], axLim[3]], 'r')
    pylab.axis(axLim)
    pylab.grid('on')
    pylab.xlabel('Plant Height (cm)', fontdict={'size':fontSize})
    pylab.ylabel('Cum. Distr.(Yc) (kg)', fontdict={'size':fontSize})
    pylab.xlim([0, 350])
    pylab.title('Height/Yc for plot %s' % (plotKey), fontdict={'size':fontSize})
    if i >= plots.__len__():
        pylab.legend(h, ['50cm threshold'], loc='upper left', bbox_to_anchor=(1.2, 1), prop={'size':fontSize})
    i += 1

    idx = height > 50
    print str(plotKey), ": "
    print "Ttl: ", str(yc.sum())
    print "Hgt>50: ", str(yc[idx].sum())
    print "Hgt>50/Plot Ttl: ", str(yc[idx].sum()/yc.sum())
    print "Hgt>50/Ttl: ", str(yc[idx].sum()/ycTtl)



#---------------------------------------------------------------------------------------------------------------
# simulate how linearly allometric models behave to scaling in canopy area / diameter

# first take a look at how much each species contributes

allSpecies = np.array([k for k, v in allometricModels.iteritems()])
allSpeciesYc = dict(zip(allSpecies, np.zeros(allSpecies.shape)))
allSpecies = np.array([k for k, v in allSpeciesYc.iteritems()])  # order can change!!!

for plotKey, plot in plots.iteritems():
    for record in plot:
        if allSpecies.__contains__(record['species']):
            if record['species'].__contains__('junceum'):
                print 'JUNCEUM'
                break
            allSpeciesYc[record['species']] += record['yc']


pylab.figure()
pylab.bar(range(0, allSpecies.size), [v for k,v in allSpeciesYc.iteritems()])
# pylab.yscale('log')
pylab.xticks(range(0, allSpecies.size), allSpecies, rotation='vertical', fontSize=fontSize)  #prop={'size':fontSize-2})
pylab.grid('on')
pylab.ylabel('Total C (kg)', fontdict={'size':fontSize})
pylab.title('Total Trial C per Species', fontdict={'size':fontSize})

#find 4 highest contributing species
idx = np.flipud(np.argsort([v for k,v in allSpeciesYc.iteritems()]))
topSpecies = allSpecies[idx[:4]]

# -------------------------------------------------------------------------------------------------------
# see how y varies with x for these species


# a hacked version of the fn above to return y for varying x
def EvalRecordCsV(allometricModels, record):
    # vars = [model['vars'] for model in allometricModels.values()]
    if not allometricModels.__contains__(record['species']):
        print record['species'], " not found"
        return 0.
    model = allometricModels[record['species']]
    x = 0.
    CD = np.mean([record['canopyLength'], record['canopyWidth']])  #make these the max valies
    CD = np.linspace(10, CD, 100)
    CA = np.pi*(CD/2)**2
    if model['vars'] == 'CA.H':
        x = CA*record['height']
    elif model['vars'] == 'CA.SL':
        x = CA*record['height']
    elif model['vars'] == 'CD':
        x = CD
    elif model['vars'] == 'CD.H':
        x = CD*record['height']
    elif model['vars'] == 'Hgt':
        x = record['height']
    else:
        print model['vars'], " unknown variable"
        return 0.
    yn = np.exp(model['ay'])*x**model['by']   # "naive"
    yc = yn*model['MB']
    if model['useWdRatio']:
        if model.__contains__('wdRatio'):
            yc = yc * model['wdRatio']
        else:
            print "WD Ratio for ", record['species'], " not found - using 1."

    return yc, x


pylab.figure()
pi = 0
for specie in topSpecies:
    model = allometricModels[specie]
    # max values
    record['canopyLength'] = 200
    record['canopyWidth'] = 200
    record['height'] = 150
    record['species'] = specie

    yc,x = EvalRecordCsV(allometricModels, record)
    pylab.subplot(2, 2, pi+1)
    pylab.plot(x, yc)
    pylab.grid()
    pylab.ylabel('Yc (kg)', fontdict={'size':fontSize})
    pylab.xlabel(allometricModels[record['species']]['vars'], fontdict={'size':fontSize})
    pylab.title(specie, fontdict={'size':fontSize})
    pi += 1




#------------------------------------------------------------------------------------------------------------------
# suspect code below
pylab.figure()
pi = 0
for specie in topSpecies:
    model = allometricModels[specie]
    canopyLengths = []
    canopyLengths = np.linspace(20,300,100)
    ycs = []
    for canopyLength in canopyLengths:  # i in range(0, 100):
        record = {}
        if False:
            record['canopyLength'] = np.random.rand(1)*200
            canopyLengths.append(record['canopyLength'])
        else:
            record['canopyLength'] = canopyLength
        record['canopyWidth'] = canopyLength*0.5
        record['height'] = 150
        record['species'] = specie

        edgeSection = record['canopyLength']*0.2
        sectionArea, ySection = CircleSectionArea(edgeSection + 0.5*record['canopyLength'], 0.5*record['canopyWidth'], edgeSection)
        record['canopyWidth'] = edgeSection*2

        # get yc for full canopy
        ycFull = EvalRecordCs(allometricModels, record)   # find c for full canopy

        ycs.append(ycFull)

    ycs = np.array(ycs)
    pylab.subplot(2, 2, pi + 1)
    pylab.plot(canopyLengths, ycs)
    pylab.plot([canopyLengths.min(), canopyLengths.max()], [ycs.min(), ycs.max()])
    pylab.grid()
    pylab.ylabel('Yc')
    pylab.xlabel('Canopy Length')
    pylab.title(specie)
    pi += 1

    if False:

        # update record for section of canopy and get yc
        sectionArea, ySection = CircleSectionArea(record['canopyLength']/2, record['canopyWidth']/2, edgeSection)
        record['canopyLength'] = record['canopyLength'] - edgeSection
        record['canopyWidth'] = ySection
        ycSectionApprox = EvalRecordCs(allometricModels, record)   #find c for approx inside area of canopy

        # get yc for actual portion of canopy
        model = allometricModels[record['species']]
        x = 0.
        CD = np.mean([record['canopyLength'], record['canopyWidth']])
        CA = sectionArea  #np.pi*(CD/2)**2
        if model['vars'] == 'CA.H':
            x = CA*record['height']
        elif model['vars'] == 'CA.SL':
            x = CA*record['height']
        elif model['vars'] == 'CD':
            x = CD
        elif model['vars'] == 'CD.H':
            x = CD**record['height']
        elif model['vars'] == 'Hgt':
            x = record['height']
        else:
            print model['vars'], " unknown variable"
        yn = model['ay']*x**model['by']   # "naive"
        Yc = yn*model['duan']
