# this is now a generic script to process any sampling data
# ----------------------------------------------------------------------------------------------------------------------
# includes woody and litter where available
# ----------------------------------------------------------------------------------------------------------------------
# Read in Allometric model parameters

allometryFileName = "C:/Data/Development/Projects/PhD GeoInformatics/Data/GEF Field Trial/AllometricModels.xlsx"
unknownSpeciesFileName = "C:/Data/Development/Projects/PhD GeoInformatics/Data/GEF Sampling/Cos_guilds_Unknown Species MV.xlsx"

# woodyFileName =  "C:/Data/Development/Projects/PhD GeoInformatics/Data/GEF Sampling/Sampling Dec 2017/GEF_Woody spp_2018.03.10_MdodaQC.xlsx"
# litterFileName = "C:/Data/Development/Projects/PhD GeoInformatics/Data/GEF Sampling/Sampling Dec 2017/GEF_Litter_2018.03.09_Cos.xlsx"

# woodyFileName = "C:/Data/Development/Projects/PhD GeoInformatics/Data/GEF Sampling/Sampling March 2018/GEF_Woody spp_INTACT_2018.04.24_Mdoda.xlsx"
# litterFileName = None

woodyFileName = "C:/Data/Development/Projects/PhD GeoInformatics/Data/GEF Sampling/GEF_Woody spp_Consolidated_ALL plots_2018.07.31_Mdoda.xlsx"
litterFileName = "C:/Data/Development/Projects/PhD GeoInformatics/Data/GEF Sampling/Litter Consolotaed ALL plots_2018.07.31.xlsx"

# woodyFileName = "C:/Data/Development/Projects/PhD GeoInformatics/Data/GEF Sampling/Sampling June 2018/GEF_SS_Woody spp_2018.07.13.Mdoda.xlsx"
# litterFileName = "C:/Data/Development/Projects/PhD GeoInformatics/Data/GEF Sampling/Sampling June 2018/GEF_Litter_SS plots_2018.07.23_Cos.xlsx"


from openpyxl import load_workbook
import numpy as np
import pylab
from scipy.stats import gaussian_kde
import collections
from csv import DictWriter
from collections import OrderedDict
import os.path
from scipy import stats as stats

fontSize = 16


def EvalRecordCs(allometricModels, record):
    # vars = [model['vars'] for model in allometricModels.values()]
    res = {'yc':0., 'area': 0., 'vol':0.}
    if not allometricModels.__contains__(record['species']):
        print record['species'], " not found"
        return res
    model = allometricModels[record['species']]
    x = 0.
    CD = np.mean([record['canopyLength'], record['canopyWidth']])
    CA = np.pi*(CD/2)**2
    res['height'] = record['height']
    res['area'] = CA
    res['vol'] = CA*record['height']   # volume of a cylinder
    # res['vol'] = (np.pi*4./3) * record['canopyLength'] * record['canopyWidth'] * record['height']/2.   # volume of an ellipse

    if model['vars'] == 'CA.H':
        x = CA*record['height']
    elif model['vars'] == 'CA.SL':
        x = CA*record['height']
    elif model['vars'] == 'CD':
        x = CD
    elif model['vars'] == 'CD.H':
        x = CD*record['height']
    elif model['vars'] == 'Hgt':
        x = record['height']
    else:
        print model['vars'], " unknown variable"
        return 0.
    yn = np.exp(model['ay'])*x**model['by']   # "naive"
    Yc = yn*model['MB']

    if model['useWdRatio']:
        if model.__contains__('wdRatio'):
            Yc = Yc * model['wdRatio']
        # else:
        #     print "WD Ratio for ", record['species'], " not found - using 1."
    res['yc'] = Yc

    return res

# def EvalRecordCs(allometricModels, record):
#     # vars = [model['vars'] for model in allometricModels.values()]
#     if not allometricModels.__contains__(record['species']):
#         print record['species'], " not found"
#         return 0.
#     model = allometricModels[record['species']]
#     x = 0.
#     CD = np.mean([record['canopyLength'], record['canopyWidth']])
#     CA = np.pi*(CD/2)**2
#     if model['vars'] == 'CA.H':
#         x = CA*record['height']
#     elif model['vars'] == 'CA.SL':
#         x = CA*record['height']
#     elif model['vars'] == 'CD':
#         x = CD
#     elif model['vars'] == 'CD.H':
#         x = CD*record['height']
#     elif model['vars'] == 'Hgt':
#         x = record['height']
#     elif model['vars'] == 'CA':
#         x = CA
#     else:
#         print model['vars'], " unknown variable"
#         return 0.
#     # if record['species'] == 'S.longispina':   # MP's model Log10 y (C (kg) = 1.1012(Log10 canopy area (m2)) - 0.2938
#     #     Yc = 10**((model['ay'] * np.log10(x)) + model['by'])
#     # else:
#     yn = np.exp(model['ay'])*x**model['by']     # "naive"
#     Yc = yn*model['MB']
#
#     if model['useWdRatio']:
#         if model.__contains__('wdRatio'):
#             Yc = Yc * model['wdRatio']
#         # else:
#         #     print "WD Ratio for ", record['species'], " not found - using 1."
#     return Yc

def EvalPlotCs(allometricModels, plot):
    plotYc = []
    for record in plot:
        Yc = EvalRecordCs(allometricModels, record)
        if Yc > 0.:
            plotYc.append(Yc)
        print record['species'], " - ", str(Yc)
    return plotYc

def MyUnknownSpeciesMap(species):
    mapSpecies = None
    if species == 'A.ferox' or species.__contains__('Aloe') or species.__contains__('A.ferox'):
        mapSpecies = 'A.speciosa'
    # if species.__contains__('unk'):
    if species.__contains__('Asparagus') or species.__contains__('hairy'):
        mapSpecies = 'A.capensis'
    if species.__contains__('Crassula') or species.__contains__('rupestris'):
        mapSpecies = 'C.ovata'
    if species.__contains__('Crassula') or species.__contains__('horns'):
        mapSpecies = 'C.perforata'
    if species == 'Euclea sp1' or species.__contains__('Euclea'):
        mapSpecies = 'E.undulata'
    if species.__contains__('Euphorbia') or species.__contains__('Euphorbia') or species.__contains__('Eurphobia'):
        mapSpecies = 'E.coerulescens'
    if species == 'Schotia sp1':
        mapSpecies = 'S.afra'
    if str.lower(species).__contains__('grewia'):
        mapSpecies = 'G.robusta'
    if species.__contains__('Lycium'):
        mapSpecies = 'L.ferocissimum'
    if species.__contains__('Ysterhout') or species.__contains__('Acacia'):  # very crude guesses
        mapSpecies = 'P.capensis'
    if species.__contains__('Gymnosporia'):  # very crude guesses
        mapSpecies = 'G.polyacantha'
    if species.__contains__('Boscia'):
        mapSpecies = 'B.oleoides'
    if species == 'A.striatus':
        mapSpecies = 'A.capensis'
    if species.__contains__('Rhigozum'):  # very crude guesses
        mapSpecies = 'R.obovatum'
    if species.__contains__('Searsia') or species == 'S.longspina' or species == 'S.longispina':
        mapSpecies = 'S.longispina'
    if species.__contains__('Crassula'):
        mapSpecies = 'C.ovata'
    if species.__contains__('Polygala'):
        mapSpecies = 'S.longispina'
    if species.__contains__('karroo'):
        mapSpecies = 'V.karoo'
    if species.__contains__('cancer'):
        mapSpecies = 'P.incana'
    if species.__contains__('Slangbossie'):
        mapSpecies = 'P.incana'
    if species.__contains__('oak'):
        mapSpecies = 'V.karoo'
    if species.__contains__('four star'):
        mapSpecies = 'S.longispina'

    return mapSpecies

def ScatterD(x, y, labels=None, class_labels=None, thumbnails=None, regress=True, xlabel=None, ylabel=None):
    if class_labels is None:
        class_labels = np.zeros(x.__len__())
    classes = np.unique(class_labels)
    colours = ['r', 'm', 'b', 'g', 'y', 'k', 'o']
    ylim = [np.min(y), np.max(y)]
    xlim = [np.min(x), np.max(x)]
    xd = np.diff(xlim)[0]
    yd = np.diff(ylim)[0]
    pylab.axis([np.min(x), np.max(x), np.min(y), np.max(y)])
    pylab.hold('on')
    ax = pylab.gca()
    handles = [0, 0, 0]

    for ci, (class_label, colour) in enumerate(zip(classes, colours[:classes.__len__()])):
        class_idx = class_labels == class_label
        if thumbnails is None:
            pylab.plot(x[class_idx], y[class_idx], colour + 'x')

        for xyi, (xx, yy) in enumerate(zip(x[class_idx], y[class_idx])):  # , np.array(plot_names)[class_idx]):
            if labels is not None:
                pylab.text(xx - .0015, yy - .0015, np.array(labels)[class_idx][xyi],
                           fontdict={'size': 9, 'color': colour, 'weight': 'bold'})

            if thumbnails is not None:
                imbuf = np.array(thumbnails)[class_idx][xyi]
                ims = 20.
                extent = [xx - xd / (2 * ims), xx + xd / (2 * ims), yy - yd / (2 * ims), yy + yd / (2 * ims)]
                pylab.imshow(imbuf[:, :, :3], extent=extent, aspect='auto')  # zorder=-1,
                handles[ci] = ax.add_patch(
                    patches.Rectangle((xx - xd / (2 * ims), yy - yd / (2 * ims)), xd / ims, yd / ims, fill=False,
                                      edgecolor=colour, linewidth=2.))
                # pylab.plot(mPixels[::step], dRawPixels[::step], color='k', marker='.', linestyle='', markersize=.5)
        if regress and classes.__len__() > 1 and False:
            (slope, intercept, r, p, stde) = stats.linregress(x[class_idx], y[class_idx])
            pylab.text(xlim[0] + xd * 0.7, ylim[0] + yd * 0.05 * (ci + 2),
                       str.format('{1}: $R^2$ = {0:.2f}', np.round(r ** 2, 2), classes[ci]),
                       fontdict={'size': 10, 'color': colour})

    if regress:
        (slope, intercept, r, p, stde) = stats.linregress(x, y)
        pylab.text((xlim[0] + xd * 0.7), (ylim[0] + yd * 0.05), str.format('$R^2$ = {0:.2f}', np.round(r ** 2, 2)),
                   fontdict={'size': 12})
        print str.format('$R^2$ = {0:.2f}', np.round(r ** 2, 2))
        print str.format('P (slope=0) = {0:.2f}', np.round(p, 3))
        print str.format('Slope = {0:.2f}', np.round(slope, 3))
        print str.format('Std error of slope = {0:.2f}', np.round(stde, 3))
        yhat = x*slope + intercept
        print str.format('RMS error = {0:.2f}', np.round(np.sqrt(np.mean((y-yhat)**2)), 3))


    if xlabel is not None:
        pylab.xlabel(xlabel, fontdict={'size': 12})
    if ylabel is not None:
        pylab.ylabel(ylabel, fontdict={'size': 12})
    # pylab.ylabel(yf)
    if classes.__len__() > 1:
        pylab.legend(handles, classes, fontsize=12)

# ---------------------------------------------------------------------------------------------------------------------
# load allometric models

wb = load_workbook(allometryFileName)
ws = wb["Allometric Models"]

first_row = ws[0 + 1]
header = []
for c in first_row:
    header.append(c.value)

allometricModels = {}
for r in ws[2:ws.max_row]:  #how to find num rows?
    if r[0].value is None:
        break
    species = str.strip(str(r[0].value[0]) + '.' + str(r[1].value))
    model = {}
    model['vars'] = r[3].value
    model['ay'] = r[6].value
    model['by'] = r[7].value
    model['MB'] = r[13].value
    model['duan'] = r[12].value
    if str(r[14].value) == 'x':
        model['useWdRatio'] = False
    else:
        model['useWdRatio'] = True
    allometricModels[species] = model
ws = None

ws = wb["Wet Dry Ratios"]
first_row = ws[0 + 1]
header = []
for c in first_row:
    header.append(c.value)

wdRatios = {}
for r in ws[2:ws.max_row]:  #how to find num rows?
    if r[0].value is None:
        break
    species = str.strip(str(r[0].value[0]) + '.' + str(r[1].value))
    model = {}
    model['WDratio'] = r[4].value
    if allometricModels.__contains__(species):
        allometricModels[species]['wdRatio'] = r[4].value
    else:
        print species, ' not found in allometric models'
    wdRatios[species] = model
ws = None

wb = None
#----------------------------------------------------------------------------------------------------------------------
# read in unknown species map
wb = load_workbook(unknownSpeciesFileName)

unknownSpeciesMap = {}
ws = wb.worksheets[0]
first_row = ws[1]
header = []
for c in first_row:
    header.append(c.value)
for r in ws[2:ws.max_row]:
    if r[2].value is None:  # no mapping yet
        continue
    unknownSpecies = str(r[0].value).strip()
    mapSpecies = str(r[2].value).strip().replace('. ', '.')
    unknownSpeciesMap[unknownSpecies] = mapSpecies
wb = None

# ----------------------------------------------------------------------------------------------------------------------
# Read in woody measurements

wb = load_workbook(woodyFileName)

# modify above map with my guesses and then, make a list of "not found" species i.e. species without allometric models, without any map
notFoundSpecies = {}
for ws in wb:
    print ws.title, ' rows: ', ws.max_row
    first_row = ws[1]
    header = []
    for c in first_row:
        header.append(c.value)

    for r in ws[2:ws.max_row]:
        if r[2].value is None:
            continue
        species = str(r[4].value).strip().replace('. ','.')
        print '.',

        if not allometricModels.__contains__(species):
            print species, " not found in allometric models"
            if not unknownSpeciesMap.__contains__(species):
                print species, " not found in guild map"
                mapSpecies = MyUnknownSpeciesMap(species)
                if mapSpecies is not None:
                    unknownSpeciesMap[species] = mapSpecies

                print species, " not found in models or maps"
                if notFoundSpecies.has_key(species):
                    notFoundSpecies[species] += 1
                else:
                    notFoundSpecies[species] = 1

# if MV:
#     outFileName = 'C:/Data/Development/Projects/PhD GeoInformatics/Data/GEF Sampling/Sampling June 2018/Unknown Species MV.csv'
# else:
#     outFileName = 'C:/Data/Development/Projects/PhD GeoInformatics/Data/GEF Sampling/Sampling June 2018/Unknown Species SS.csv'
outFileName = str.format('{0}/{1} - Unknown Species.csv', os.path.dirname(woodyFileName), os.path.splitext(os.path.basename(woodyFileName))[0])

import csv
with open(outFileName, 'wb') as outfile:
    writer = csv.writer(outfile, delimiter=',',
                            quotechar='/',  quoting=csv.QUOTE_MINIMAL)
    writer.writerow(['Species', 'N'])
    for species, N in notFoundSpecies.iteritems():
        writer.writerow([species, N])

print '------------------------------------------------------'

#
# for ws in wb:
#     print ws.title, ' rows: ', ws.max_row
#     header = []
#
#     for r in ws[2:ws.max_row]:
#         if r[0].value is None:
#             break
#         species = str(r[2].value).strip().replace('. ','.')
#
#         if not allometricModels.__contains__(species):
#             print species, " not found"
#             if notFoundSpecies.has_key(species):
#                 notFoundSpecies[species] += 1
#             else:
#                 notFoundSpecies[species] = 1


# read in allometry and find woody cs
nestedPlots = {}
for ws in wb:
    print ws.title, ' rows: ', ws.max_row
    first_row = ws[1]
    header = []
    for c in first_row:
        header.append(c.value)

    plots = collections.OrderedDict()
    for r in ws[2:ws.max_row]:
        if r is None or r[2].value is None:
            continue
        # print  r[2].value
        record = OrderedDict()
        species = str(r[4].value).strip().replace('. ','.')
        id = str(r[2].value).strip()
        dashLoc = str(id).find('-')
        if dashLoc < 0:
            dashLoc = 2
        id = id.replace('-', '')
        idNum = np.int32(id[dashLoc:])  # get rid of leading zeros
        id = '%s%d' % (id[:dashLoc], idNum)

        record['ID'] = id
        record['degrClass'] = str(r[3].value)
        record['orig. species'] = str(species).strip()
        record['canopyWidth'] = r[5].value
        record['canopyLength'] = r[6].value
        record['height'] = r[7].value

        if not allometricModels.__contains__(species):
            if not unknownSpeciesMap.__contains__(species):
                print species, " not found, no subs"
            else:
                species = unknownSpeciesMap[species]
        record['species'] = str(species).strip()

        # zerp empty records
        fields = ['canopyWidth', 'canopyLength', 'height']
        for f in fields:
            if record[f] is None:
                record[f] = 0

        if r.__len__() > 8:
            record['bsd'] = str(r[8].value)
        else:
            record['bsd'] = ""
        res = EvalRecordCs(allometricModels, record)
        record['yc'] = res['yc']
        record['area'] = res['area']
        record['vol'] = res['vol']

        if plots.has_key(id):
            plots[id].append(record)
        else:
            plots[id] = [record]
    nestedPlots[ws.title] = plots

    # outFileName = 'C:/Data/Development/Projects/PhD GeoInformatics/Code/Results/Baviaans2017FieldTrialAnalysis/%s - Woody.csv' % (ws.title)
    # with open(outFileName,'wb') as outfile:
    #     writer = DictWriter(outfile, plot[0].keys())
    #     writer.writeheader()
    #     writer.writerows(plot)
    # if MV:
    #     outFileName = 'C:/Data/Development/Projects/PhD GeoInformatics/Data/GEF Sampling/Sampling June 2018/%s - MV Woody.csv' % (ws.title)
    # else:
    #     outFileName = 'C:/Data/Development/Projects/PhD GeoInformatics/Data/GEF Sampling/Sampling June 2018/%s - SS Woody.csv' % (ws.title)
    # outFileName = str.format('{0} - {1} - Woody.csv', os.path.dirname(woodyFileName),
    #                          os.path.splitext(os.path.basename(woodyFileName))[0])
    outFileName = str.format('{0}/{1} - {2} - Woody.csv', os.path.dirname(woodyFileName),
                             os.path.splitext(os.path.basename(woodyFileName))[0], ws.title)

    with open(outFileName,'wb') as outfile:
        writer = DictWriter(outfile, plots.values()[0][0].keys())
        writer.writeheader()
        for plot in plots.values():
            writer.writerows(plot)

wb = None
#----------------------------------------------------------------------------------------------------------------
# read in litter data
litterDict = {}
if litterFileName is not None:
    wb = load_workbook(litterFileName, data_only=True)

    litterDict = {}
    ws = wb['Litter summary']
    for r in ws[2:ws.max_row]:
        id = str(r[0].value).strip()
        id = id.replace('-0','')
        id = id.replace('-','')
        if id == '' or id is None or id == 'None' or r[1].value == 0:
            print 'No ID, continue'
            continue
        else:
            print id,
        dryWeight = r[2].value
        if litterDict.has_key(id):
            litterDict[id]['dryWeight'] += dryWeight
        else:
            litterDict[id] = {'dryWeight': dryWeight}
    wb = None
else:
    litterDict = {}
    for id, plot in nestedPlots['5x5m'].iteritems():
        litterDict[id] = {'dryWeight': -1.}

#---------------------------------------------------------------------------------------------------------------
# extrapolate 5x5m subplots to full size
# (separate contributions from 5x5m plants ><50cm high, extrap <50 to full size, add >50 to full size)
pairedPlots = {}

for key, plots in nestedPlots.iteritems():
    print key
    dashLoc = str(key).find('-')
    if dashLoc < 0:
        dashLoc = str(key).find('_')
    id = key[:dashLoc]
    plotSizeStr = key[dashLoc+1:]
    xLoc = str(plotSizeStr).find('x')
    plotSize = np.int32(plotSizeStr[:xLoc])
    if not pairedPlots.has_key(id):
        pairedPlots[id] = {}
    if plotSize == 5:
        pairedPlots[id]['innerPlots'] = plots
        pairedPlots[id]['innerPlotSize'] = 5
    else:
        pairedPlots[id]['outerPlots'] = plots
        pairedPlots[id]['outerPlotSize'] = plotSize

summaryPlots = {}
for pairedId, pairedPlot in pairedPlots.iteritems():
    outerPlots = pairedPlot['outerPlots']
    outerSize = pairedPlot['outerPlotSize']
    if pairedPlot.has_key('innerPlotSize'):
        subPlots = pairedPlot['innerPlots']
    else:
        subPlots = None

    for id, outerPlot in outerPlots.iteritems():
        summaryPlots[id] = {}
        if subPlots is not None:
            subPlot = subPlots[id]
            subHeight = np.array([r['height'] for r in subPlot])
            subYc = np.array([r['yc'] for r in subPlot])
            subArea = np.array([r['area'] for r in subPlot])
            subVol = np.array([r['vol'] for r in subPlot])
            smallIdx = subHeight < 50
            # outerHeight = np.array([r['height'] for r in outerPlot])
            outerYc = np.array([r['yc'] for r in outerPlot])
            outerArea = np.array([r['area'] for r in outerPlot])
            outerVol = np.array([r['vol'] for r in outerPlot])
            outerHeight = np.array([r['height'] for r in outerPlot])
            summaryYc = ((outerSize/5.)**2) * subYc[smallIdx].sum() + subYc[np.logical_not(smallIdx)].sum() + outerYc.sum()
            summaryPlots[id]['N'] = ((outerSize / 5.) ** 2) * smallIdx.sum() + (subYc.__len__() - smallIdx.sum()) + outerYc.__len__()
            summaryArea = ((outerSize/5.)**2) * subArea[smallIdx].sum() + subArea[np.logical_not(smallIdx)].sum() + outerArea.sum()
            summaryVol = ((outerSize/5.)**2) * subVol[smallIdx].sum() + subVol[np.logical_not(smallIdx)].sum() + outerVol.sum()
            summaryHeights = np.array(np.int32((outerSize/5.)**2) * (subHeight[smallIdx].tolist()) + subHeight[np.logical_not(smallIdx)].tolist() + outerHeight.tolist())
        else:
            outerYc = np.array([r['yc'] for r in outerPlot])
            outerArea = np.array([r['area'] for r in outerPlot])
            outerVol = np.array([r['vol'] for r in outerPlot])
            summaryHeights = np.array([r['height'] for r in outerPlot])
            summaryYc = outerYc.sum()
            summaryPlots[id]['N'] = outerYc.__len__()
            summaryArea = outerArea.sum()
            summaryVol = outerVol.sum()

        summaryPlots[id]['ID'] = id
        summaryPlots[id]['Yc'] = summaryYc
        summaryPlots[id]['Size'] = outerSize
        summaryPlots[id]['Degr. Class'] = outerPlot[0]['degrClass']
        summaryPlots[id]['YcHa'] = (100.**2) * summaryYc/(outerSize**2)
        summaryPlots[id]['Area'] = summaryArea
        summaryPlots[id]['AreaHa'] = (100.**2) * summaryArea/(outerSize**2)
        summaryPlots[id]['Vol'] = summaryVol
        summaryPlots[id]['VolHa'] = (100.**2) * summaryVol/(outerSize**2)
        summaryPlots[id]['MaxHeight'] = summaryHeights.max()
        summaryPlots[id]['MeanHeight'] = summaryHeights.mean()
        summaryPlots[id]['MedianHeight'] = np.median(summaryHeights)
        summaryPlots[id]['SumHeight'] = summaryHeights.sum()
        summaryPlots[id]['SumHeightHa'] = (100.**2) * summaryHeights.sum()/(outerSize**2)

        if litterDict.has_key(id) and litterDict[id]['dryWeight'] > 0.:
            summaryPlots[id]['Litter'] = litterDict[id]['dryWeight']/1000  # g to kg
            # the litter quadrats are uniform across all plot sizes.  the dry weight is converted to carbon using the factor 0.48 (according to Cos)
            # or 0.37 (according to James / CDM)
            # current the trial plots have no litter data (INT* and TCH*)
            # NB note, there is a factor of 44/12 in the CDM equations that I don't quite understand.  I think it is a conversion from C weight to
            # CO2 weight (i.e. a chemical thing that represents not how much carbon is stored by how much C02 was captured out the atmosphere).
            # we should make sure that we are using the same units in the woody and litter C i.e. we need to check what units Marius eq give
            # summaryPlots[id]['LitterHa'] = summaryPlots[id]['Litter'] * 0.48 * (100.**2) / (4 * (0.5**2))
            summaryPlots[id]['LitterHa'] = summaryPlots[id]['Litter'] * 0.37 * (100. ** 2) / (4 * (0.5 ** 2))
            summaryPlots[id]['AgbHa'] = summaryPlots[id]['LitterHa'] + summaryPlots[id]['YcHa']
        else:
            summaryPlots[id]['Litter'] = -1.
            summaryPlots[id]['LitterHa'] = -1.
            summaryPlots[id]['AgbHa'] = summaryPlots[id]['YcHa']


# write out summary ground truth for each plot
# if MV:
#     outFileName = 'C:/Data/Development/Projects/PhD GeoInformatics/Data/GEF Sampling/Sampling June 2018/Summary - MV Woody.csv'
# else:
#     outFileName = 'C:/Data/Development/Projects/PhD GeoInformatics/Data/GEF Sampling/Sampling June 2018/Summary - SS Woody.csv'
outFileName = str.format('{0}/{1} - Summary Woody & Litter.csv', os.path.dirname(woodyFileName),
                         os.path.splitext(os.path.basename(woodyFileName))[0])

with open(outFileName, 'wb') as outfile:
    writer = DictWriter(outfile, summaryPlots.values()[0].keys())
    writer.writeheader()
    writer.writerows(summaryPlots.values())

#---------------------------------------------------------------------------------------------------------------
# get stats and std errors per stratum

ids = np.array([plot['ID'] for plot in summaryPlots.values()])
ycHas = np.array([plot['YcHa'] for plot in summaryPlots.values()])
agbHas = np.array([plot['AgbHa'] for plot in summaryPlots.values()])
litterHas = np.array([plot['LitterHa'] for plot in summaryPlots.values()])
degrClasses = np.array([plot['Degr. Class'] for plot in summaryPlots.values()])

statsDict={}
for degrClass in np.unique(degrClasses):
    classIdx = degrClass == degrClasses
    plotStats = {}
    plotStats['MeanYcHa'] = ycHas[classIdx].mean()
    plotStats['StdYcHa'] = ycHas[classIdx].std()
    plotStats['SeYcHa'] = ycHas[classIdx].std()/np.sqrt(classIdx.sum())
    plotStats['Se:MeanYcHa'] = 100*plotStats['SeYcHa']/plotStats['MeanYcHa']

    plotStats['MeanAgbHa'] = agbHas[classIdx].mean()
    plotStats['StdAgbHa'] = agbHas[classIdx].std()
    plotStats['SeAgbHa'] = agbHas[classIdx].std() / np.sqrt(classIdx.sum())
    plotStats['Se:MeanAgbHa'] = 100*plotStats['SeAgbHa']/plotStats['MeanAgbHa']

    plotStats['MeanLitterHa'] = litterHas[classIdx].mean()
    plotStats['StdLitterHa'] = litterHas[classIdx].std()
    plotStats['SeLitterHa'] = litterHas[classIdx].std() / np.sqrt(classIdx.sum())
    plotStats['Se:MeanLitterHa'] = 100 * plotStats['SeLitterHa'] / plotStats['MeanLitterHa']

    statsDict[degrClass] = plotStats

    print 'Degradation Class: %s , N: %d'%(degrClass, classIdx.sum())
    print '--------------------------------------'
    print 'Mean (Std) Woody C (kg/ha): %.0f (%.0f)'%(plotStats['MeanYcHa'], plotStats['StdYcHa'])
    print 'Mean (Std) Litter C (kg/ha): %.0f (%.0f)' % (plotStats['MeanLitterHa'], plotStats['StdLitterHa'])
    print 'Mean (Std) AGB C (kg/ha): %.0f (%.0f)' % (plotStats['MeanAgbHa'], plotStats['StdAgbHa'])
    print 'SE/Mean Woody (%%): %.2f' % (plotStats['Se:MeanYcHa'])
    print 'SE/Mean Litter (%%): %.2f' % (plotStats['Se:MeanLitterHa'])
    print 'SE/Mean AGB (%%): %.2f' % (plotStats['Se:MeanAgbHa'])
    print ''

severeIdx = degrClasses == 'Severe'
svfIdx = np.array([str(id).startswith('SS') for id in ids])
tchIdx = (~svfIdx) & severeIdx & (litterHas > 0)
farmNames = ['Sewefontein', 'Tchnuganu']
farmDict = {}
for idx, farmName in zip([svfIdx, tchIdx], farmNames):
    plotStats = {}
    plotStats['MeanYcHa'] = ycHas[idx].mean()
    plotStats['StdYcHa'] = ycHas[idx].std()
    plotStats['SeYcHa'] = ycHas[idx].std()/np.sqrt(idx.sum())
    plotStats['Se:MeanYcHa'] = 100*plotStats['SeYcHa']/plotStats['MeanYcHa']

    plotStats['MeanAgbHa'] = agbHas[idx].mean()
    plotStats['StdAgbHa'] = agbHas[idx].std()
    plotStats['SeAgbHa'] = agbHas[idx].std() / np.sqrt(idx.sum())
    plotStats['Se:MeanAgbHa'] = 100*plotStats['SeAgbHa']/plotStats['MeanAgbHa']

    plotStats['MeanLitterHa'] = litterHas[idx].mean()
    plotStats['StdLitterHa'] = litterHas[idx].std()
    plotStats['SeLitterHa'] = litterHas[idx].std() / np.sqrt(idx.sum())
    plotStats['Se:MeanLitterHa'] = 100 * plotStats['SeLitterHa'] / plotStats['MeanLitterHa']

    farmDict[farmName] = plotStats

    print 'Farm: %s' % (farmName)
    print '--------------------------------------'
    print 'Mean (Std) Woody C (kg/ha): %.0f (%.0f)' % (plotStats['MeanYcHa'], plotStats['StdYcHa'])
    print 'Mean (Std) Litter C (kg/ha): %.0f (%.0f)' % (plotStats['MeanLitterHa'], plotStats['StdLitterHa'])
    print 'Mean (Std) AGB C (kg/ha): %.0f (%.0f)' % (plotStats['MeanAgbHa'], plotStats['StdAgbHa'])
    print 'SE/Mean Woody (%%): %.2f' % (plotStats['Se:MeanYcHa'])
    print 'SE/Mean Litter (%%): %.2f' % (plotStats['Se:MeanLitterHa'])
    print 'SE/Mean AGB (%%): %.2f' % (plotStats['Se:MeanAgbHa'])
    print ''

# statsDict['All'] = plotStats

#-----------------------------------------------------------------------------------------------------------------
#  look at correlation between yc / agb and height/area/vol

ycHa = np.array([plot['YcHa'] for plot in summaryPlots.values()])
agbHa = np.array([plot['AgbHa'] for plot in summaryPlots.values()])
maxHeight = np.array([plot['MaxHeight'] for plot in summaryPlots.values()])
meanHeight = np.array([plot['MeanHeight'] for plot in summaryPlots.values()])
medianHeight = np.array([plot['MedianHeight'] for plot in summaryPlots.values()])
sumHeightHa = np.array([plot['SumHeightHa'] for plot in summaryPlots.values()])
areaHa = np.array([plot['AreaHa'] for plot in summaryPlots.values()])
volHa = np.array([plot['VolHa'] for plot in summaryPlots.values()])
n = np.array([plot['N'] for plot in summaryPlots.values()])
id  = np.array([plot['ID'] for plot in summaryPlots.values()])

pylab.close('all')
fig = pylab.figure()
ScatterD(volHa/1.e6, ycHa, regress=True, labels=ids, xlabel='vol (m3/ha)', ylabel='yc (t/ha)')
pylab.grid()

fig = pylab.figure()
ScatterD((volHa**1)/1.e6, agbHa, regress=True, labels=ids, xlabel='vol (m3/ha)', ylabel='AGB (t/ha)')
pylab.grid()

fig = pylab.figure()
ScatterD(areaHa, ycHa, regress=True, xlabel='area (m2/ha)', ylabel='yc (t/ha)')
pylab.grid()

fig = pylab.figure()
ScatterD(areaHa*meanHeight/1e6, ycHa, regress=True, xlabel='area*mean(height) (m3/ha)', ylabel='yc (t/ha)')
pylab.grid()

fig = pylab.figure()
ScatterD(areaHa*(meanHeight**.5)/1e6, agbHa, regress=True, xlabel='area*sqrt(mean(height)) (m3/ha)', ylabel='AGB (t/ha)')
pylab.grid()

fig = pylab.figure()
ScatterD(np.sqrt(sumHeightHa/1.e2), ycHa, regress=True, xlabel='sum(height) (m/ha)', ylabel='yc (t/ha)')
pylab.grid()

#-----------------------------------------------------------------------------------------------------------------
#  look at height & yc distribution for 5x5m plots

plots = nestedPlots['5x5m']
# vars = [model['vars'] for model in allometricModels.values()]
# print np.unique(vars)
pylab.close('all')
i = 1
ycTtl = 0.
pylab.figure()
plotSummary = []
for plotKey, plot in plots.iteritems():
    yc = np.array([record['yc'] for record in plot])
    height = np.float64([record['height'] for record in plot])
    ycTtl += yc.sum()
    plotSummary.append({'ID': plotKey.replace('-','_'), 'Yc': yc.sum(), 'N': yc.__len__()})

    kde = gaussian_kde(height)  #, bw_method=bandwidth / height.std(ddof=1))
    heightGrid = np.linspace(0, 300, 100)
    heightKde = kde.evaluate(heightGrid)
    pylab.subplot(4, 5, i)
    pylab.plot(heightGrid, heightKde)
    axLim = pylab.axis()
    h = pylab.plot([50, 50], [0, heightKde.max()], 'r')
    pylab.grid('on')
    pylab.xlabel('Plant Height (cm)', fontdict={'size':fontSize})
    pylab.ylabel('Prob.(height)', fontdict={'size':fontSize})
    pylab.title('Height dist. for %s' % (plotKey), fontdict={'size':fontSize})
    if i >= plots.__len__():
        pylab.legend(h, ['50cm threshold'], loc='upper left', bbox_to_anchor=(1.2, 1), prop={'size':fontSize})
    pylab.axis([axLim[0], axLim[1], 0, heightKde.max()])
    i += 1

i = 1
pylab.figure()
for plotKey, plot in plots.iteritems():
    yc = np.array([record['yc'] for record in plot])
    height = np.float64([record['height'] for record in plot])
    idx = np.argsort(height)

    ycCumSum = np.cumsum(yc[idx])
    pylab.subplot(4, 5, i)
    pylab.plot(height[idx], ycCumSum)
    axLim = pylab.axis()
    h = pylab.plot([50, 50], [axLim[2], axLim[3]], 'r')
    pylab.axis(axLim)
    pylab.grid('on')
    pylab.xlabel('Plant Height (cm)', fontdict={'size':fontSize})
    pylab.ylabel('Cum. Distr.(Yc) (kg)', fontdict={'size':fontSize})
    pylab.xlim([0, 350])
    pylab.title('Height/Yc for plot %s' % (plotKey), fontdict={'size':fontSize})
    if i >= plots.__len__():
        pylab.legend(h, ['50cm threshold'], loc='upper left', bbox_to_anchor=(1.2, 1), prop={'size':fontSize})
    i += 1

    idx = height > 50
    print str(plotKey), ": "
    print "Ttl: ", str(yc.sum())
    print "Hgt>50: ", str(yc[idx].sum())
    print "Hgt>50/Plot Ttl: ", str(yc[idx].sum()/yc.sum())
    print "Hgt>50/Ttl: ", str(yc[idx].sum()/ycTtl)



#---------------------------------------------------------------------------------------------------------------
# simulate how linearly allometric models behave to scaling in canopy area / diameter

# first take a look at how much each species contributes

allSpecies = np.array([k for k, v in allometricModels.iteritems()])
allSpeciesYc = dict(zip(allSpecies, np.zeros(allSpecies.shape)))
allSpecies = np.array([k for k, v in allSpeciesYc.iteritems()])  # order can change!!!

for plotKey, plot in plots.iteritems():
    for record in plot:
        if allSpecies.__contains__(record['species']):
            if record['species'].__contains__('junceum'):
                print 'JUNCEUM'
                break
            allSpeciesYc[record['species']] += record['yc']


pylab.figure()
pylab.bar(range(0, allSpecies.size), [v for k,v in allSpeciesYc.iteritems()])
# pylab.yscale('log')
pylab.xticks(range(0, allSpecies.size), allSpecies, rotation='vertical', fontSize=fontSize)  #prop={'size':fontSize-2})
pylab.grid('on')
pylab.ylabel('Total C (kg)', fontdict={'size':fontSize})
pylab.title('Total Trial C per Species', fontdict={'size':fontSize})

#find 4 highest contributing species
idx = np.flipud(np.argsort([v for k,v in allSpeciesYc.iteritems()]))
topSpecies = allSpecies[idx[:4]]

# -------------------------------------------------------------------------------------------------------
# see how y varies with x for these species


# a hacked version of the fn above to return y for varying x
def EvalRecordCsV(allometricModels, record):
    # vars = [model['vars'] for model in allometricModels.values()]
    if not allometricModels.__contains__(record['species']):
        print record['species'], " not found"
        return 0.
    model = allometricModels[record['species']]
    x = 0.
    CD = np.mean([record['canopyLength'], record['canopyWidth']])  #make these the max valies
    CD = np.linspace(10, CD, 100)
    CA = np.pi*(CD/2)**2
    if model['vars'] == 'CA.H':
        x = CA*record['height']
    elif model['vars'] == 'CA.SL':
        x = CA*record['height']
    elif model['vars'] == 'CD':
        x = CD
    elif model['vars'] == 'CD.H':
        x = CD*record['height']
    elif model['vars'] == 'Hgt':
        x = record['height']
    else:
        print model['vars'], " unknown variable"
        return 0.
    yn = np.exp(model['ay'])*x**model['by']   # "naive"
    yc = yn*model['MB']
    if model['useWdRatio']:
        if model.__contains__('wdRatio'):
            yc = yc * model['wdRatio']
        else:
            print "WD Ratio for ", record['species'], " not found - using 1."

    return yc, x


pylab.figure()
pi = 0
for specie in topSpecies:
    model = allometricModels[specie]
    # max values
    record['canopyLength'] = 200
    record['canopyWidth'] = 200
    record['height'] = 150
    record['species'] = specie

    yc,x = EvalRecordCsV(allometricModels, record)
    pylab.subplot(2, 2, pi+1)
    pylab.plot(x, yc)
    pylab.grid()
    pylab.ylabel('Yc (kg)', fontdict={'size':fontSize})
    pylab.xlabel(allometricModels[record['species']]['vars'], fontdict={'size':fontSize})
    pylab.title(specie, fontdict={'size':fontSize})
    pi += 1




#------------------------------------------------------------------------------------------------------------------
# suspect code below
pylab.figure()
pi = 0
for specie in topSpecies:
    model = allometricModels[specie]
    canopyLengths = []
    canopyLengths = np.linspace(20,300,100)
    ycs = []
    for canopyLength in canopyLengths:  # i in range(0, 100):
        record = {}
        if False:
            record['canopyLength'] = np.random.rand(1)*200
            canopyLengths.append(record['canopyLength'])
        else:
            record['canopyLength'] = canopyLength
        record['canopyWidth'] = canopyLength*0.5
        record['height'] = 150
        record['species'] = specie

        edgeSection = record['canopyLength']*0.2
        sectionArea, ySection = CircleSectionArea(edgeSection + 0.5*record['canopyLength'], 0.5*record['canopyWidth'], edgeSection)
        record['canopyWidth'] = edgeSection*2

        # get yc for full canopy
        ycFull = EvalRecordCs(allometricModels, record)   # find c for full canopy

        ycs.append(ycFull)

    ycs = np.array(ycs)
    pylab.subplot(2, 2, pi + 1)
    pylab.plot(canopyLengths, ycs)
    pylab.plot([canopyLengths.min(), canopyLengths.max()], [ycs.min(), ycs.max()])
    pylab.grid()
    pylab.ylabel('Yc')
    pylab.xlabel('Canopy Length')
    pylab.title(specie)
    pi += 1

    if False:

        # update record for section of canopy and get yc
        sectionArea, ySection = CircleSectionArea(record['canopyLength']/2, record['canopyWidth']/2, edgeSection)
        record['canopyLength'] = record['canopyLength'] - edgeSection
        record['canopyWidth'] = ySection
        ycSectionApprox = EvalRecordCs(allometricModels, record)   #find c for approx inside area of canopy

        # get yc for actual portion of canopy
        model = allometricModels[record['species']]
        x = 0.
        CD = np.mean([record['canopyLength'], record['canopyWidth']])
        CA = sectionArea  #np.pi*(CD/2)**2
        if model['vars'] == 'CA.H':
            x = CA*record['height']
        elif model['vars'] == 'CA.SL':
            x = CA*record['height']
        elif model['vars'] == 'CD':
            x = CD
        elif model['vars'] == 'CD.H':
            x = CD**record['height']
        elif model['vars'] == 'Hgt':
            x = record['height']
        else:
            print model['vars'], " unknown variable"
        yn = model['ay']*x**model['by']   # "naive"
        Yc = yn*model['duan']
